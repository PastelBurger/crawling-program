
import requests
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup as bs
from pandas import DataFrame
from selenium import webdriver
import time
import openpyxl
import datetime

# 함수

def func_to_date(list) :
    time_format = "%Y-%m-%d"
    ss=[]
    for tag in list :
        ss.append ( datetime.datetime.strptime(tag, time_format ) )

    return ss

def func_create_Condition(list) :
    C_time = datetime.datetime.now ()
    for tag in list:
        if C_time <= tag :
            Condition.append("공고중")
        else:
            Condition.append("공고종료")
    return Condition



# <IP_NAVI_사업공고&입찰공고------------------------------------------------------------------------------------시작 1>


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options


# 브라우저 꺼짐 방지 옵션

chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=chrome_options)

# 웹페이지 해당 주소 이동

driver.implicitly_wait(3)
driver.get('http://www.ip-navi.or.kr/ipnavi/board/boardList.navi?boardCode=B00001')
time.sleep(1)

page = driver.page_source
soup = bs(page, "html.parser")


def function_IP_NAVI_사업공고() :

    page = driver.page_source
    soup = bs(page, "html.parser")
    # 헤드라인
    a = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr> td.align_L ' )
    s1 = []
    for tag in a :
        s1.append ( tag.text )
    s2 = []
    for tag in s1 :
        s2.append (
            tag.replace ( '\n' , '' ).replace ( '\t' , '' ).replace ( '  ' , '' ).replace ( '\xa0' , '' ).replace ( '\r' ,'' ) )
    a = pd.DataFrame ( s2 )
    # 링크
    s1 = [element['onclick'] for element in soup.find_all('tr', attrs={'onclick': True})]
    s1 = pd.DataFrame ( s1 )
    v_split = s1[0].str.split ( '(' )
    s1 = pd.DataFrame ( v_split.str.get ( 1 ) )
    v_split = s1[0].str.split ( ',' )
    s1 = pd.DataFrame ( v_split.str.get ( 1 ) )
    s2 = []
    for tag in s1[0] :
        s2.append ( tag[2 :-2] )
    s3=[]
    for tag in s2:
        s3.append ( "https://www.ip-navi.or.kr/ipnavi/board/boardDetail.navi?boardCode=B00001&boardSeq="+tag+"&currentPageNo=1")
    b = pd.DataFrame ( s3 )
    b.rename(columns={ 0: 'b'}, inplace=True)
    # Uploader
    s1 = []
    for tag in range ( len ( a ) ) :
        s1.append ( "IP_NAVI_사업공고" )
    c = pd.DataFrame ( s1 )
    # 날짜
    d = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr> td:nth-child(4)' )
    d = pd.DataFrame ( d )
    s1 = []
    time_format = "%Y-%m-%d"
    for tag in d[0] :
        s1.append ( datetime.datetime.strptime ( tag , time_format ) )
    d = pd.DataFrame ( s1 )
    d.rename(columns={ 0: 'd'}, inplace=True)
    # 조회수
    e = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr > td.views' )
    s1 = []
    for tag in e :
        s1.append ( tag.text )
    s2 = []
    for tag in s1 :
        s2.append ( int(tag) )
    e = pd.DataFrame ( s2 )
    # 분류
    s1 = []
    for tag in range ( len ( a ) ) :
        s1.append ( "None" )
    f = pd.DataFrame ( s1 )
    f.rename(columns={ 0: 'f'}, inplace=True)
    # 번호
    g = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr > td:nth-child(1)' )
    g = pd.DataFrame ( g )
    # 헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
    # Site :  c , Number: g, Headline: a, Link : b, Uploader : f, Date1 : d, Deadline : f, Condition : f
    df = []
    df = pd.merge ( c , g , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , a , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , b , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , f , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , f , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , f , left_index=True , right_index=True , how='outer' )

    return df

# 1번 ~ 2번
IP_NAVI_사업공고 = function_IP_NAVI_사업공고() # 1번
time.sleep(1)
element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[1]')
driver.execute_script("arguments[0].click();", element)
time.sleep(1)
a = function_IP_NAVI_사업공고()
IP_NAVI_사업공고=pd.concat([IP_NAVI_사업공고,a]) #2번 합치기
# 3번
element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[4]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(3)
time.sleep(1)
a = function_IP_NAVI_사업공고()
IP_NAVI_사업공고=pd.concat([IP_NAVI_사업공고,a])

# 4번
element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[5]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(3)
time.sleep(1)
a = function_IP_NAVI_사업공고()
IP_NAVI_사업공고=pd.concat([IP_NAVI_사업공고,a])
# 5번
element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[6]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(3)
time.sleep(1)
a = function_IP_NAVI_사업공고()
IP_NAVI_사업공고=pd.concat([IP_NAVI_사업공고,a])


IP_NAVI_사업공고.columns=["Site","Number",'Headline',"Link",'Uploader',"Date1", 'Deadline', 'Condition']



# < IP_NAVI_사업공고&입찰공고 - -----------------------------------------------------------------------------------완료 >


# < RIPC_입찰공고 - -----------------------------------------------------------------------------------시작 2>

chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=chrome_options)

driver.implicitly_wait(3)
driver.get('https://pms.ripc.org/pms/biz/applicant/notice/list.do')
time.sleep(1)

page = driver.page_source
soup = bs(page, "html.parser")
Data1 = soup.select(' div.table_area > table > tbody > tr > td')

def function_RIPC_사업공고() :

    page = driver.page_source
    soup = bs(page, "html.parser")
    Data1=soup.select(' div.table_area > table > tbody > tr > td')

    # 링크 생성
    d = [tag for tag in Data1 if Data1.index(tag) % 6 == 3]
    link = []
    for tag in d:
        link.append("https://pms.ripc.org" + tag['onclick'][15:-1])
    link = pd.DataFrame(link)
    link.rename(columns={ 0: 'link'}, inplace=True)
    # 다른항목 생성
    a = pd.DataFrame(tag.text for tag in Data1 if Data1.index(tag) % 6==0)
    a.rename(columns={ 0: 'a'}, inplace=True)
    b = pd.DataFrame(tag.text for tag in Data1 if Data1.index(tag) % 6==1)
    b.rename(columns={ 0: 'b'}, inplace=True)
    c = pd.DataFrame(tag.text for tag in Data1 if Data1.index(tag) % 6==2)
    c.rename(columns={ 0: 'c'}, inplace=True)
    d = pd.DataFrame(tag.text for tag in Data1 if Data1.index(tag) % 6==3)
    d.rename(columns={ 0: 'd'}, inplace=True)
    e = pd.DataFrame(tag.text for tag in Data1 if Data1.index(tag) % 6==4)
    e.rename(columns={ 0: 'e'}, inplace=True)
    f = pd.DataFrame(tag.text for tag in Data1 if Data1.index(tag) % 6==5)
    f.rename(columns={ 0: 'f'}, inplace=True)
    # 날짜나누기
    e.columns = ['Date']
    v_split = e.Date.str.split('~')
    v_Date1 = pd.DataFrame(v_split.str.get(0))
    v_Date2 = pd.DataFrame(v_split.str.get(1))
    time_format = "%Y-%m-%d %H:%M "
    s1 = []
    s2 = []
    time_format = "%Y-%m-%d %H:%M "
    for tag in v_Date1["Date"] :
        s1.append(tag.replace('\n','').replace('\t',''))
    s1= pd.DataFrame ( s1 )
    for tag in s1[0]:
        s2.append ( datetime.datetime.strptime ( tag , time_format ) )
    v_Date1=pd.DataFrame(s2)
    v_Date1.rename(columns={ 0: 'v_Date1'}, inplace=True)
    s1 = []
    s2 = []
    time_format = " %Y-%m-%d %H:%M"
    for tag in v_Date2["Date"] :
        s1.append(tag.replace('\n','').replace('\t',''))
    s1 = pd.DataFrame ( s1 )
    for tag in s1[0]:
        s2.append ( datetime.datetime.strptime ( tag , time_format ) )
    v_Date2=pd.DataFrame(s2)
    v_Date2.rename(columns={ 0: 'v_Date2'}, inplace=True)
    df= pd.merge(a,b, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,c, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,d, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,v_Date1, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,v_Date2, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,f, left_index=True, right_index=True, how='outer')
    df = pd.merge(df, link, left_index=True, right_index=True, how='outer')
    df['Site'] = 'IP_RIPC_입찰공고'
    return df

# 1번 ~ 2번
RIPC_사업공고 = function_RIPC_사업공고() # 1번
driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[1]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_RIPC_사업공고()
RIPC_사업공고=pd.concat([RIPC_사업공고,a]) #2번 합치기
# 3번
driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[4]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_RIPC_사업공고()
RIPC_사업공고=pd.concat([RIPC_사업공고,a])

# 4번
driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[5]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_RIPC_사업공고()
RIPC_사업공고=pd.concat([RIPC_사업공고,a])
# 5번
driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[6]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_RIPC_사업공고()
RIPC_사업공고=pd.concat([RIPC_사업공고,a])
# 6번
driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[7]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_RIPC_사업공고()
RIPC_사업공고=pd.concat([RIPC_사업공고,a])
# 7번
driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[8]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_RIPC_사업공고()
RIPC_사업공고=pd.concat([RIPC_사업공고,a])
# 8번
driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[9]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_RIPC_사업공고()
RIPC_사업공고=pd.concat([RIPC_사업공고,a])
# 9번
driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[9]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_RIPC_사업공고()
RIPC_사업공고=pd.concat([RIPC_사업공고,a])
# 10번
driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[9]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_RIPC_사업공고()
RIPC_사업공고=pd.concat([RIPC_사업공고,a])

RIPC_사업공고.columns=['AA', 'Uploader', "Number", "Headline", "Date1", "Deadline", 'Condition', "Link","Site"]
RIPC_사업공고=RIPC_사업공고[["Site","Number",'Headline',"Link",'Uploader',"Date1", 'Deadline', 'Condition' ]]


# End_result.to_excel('./Test.xlsx', sheet_name='Sheet1')

# < RIPC_입찰공고 - -----------------------------------------------------------------------------------완료 >

# < 전략개발원_사업공고 - -----------------------------------------------------------------------------------시작 3>

url="https://biz.kista.re.kr/ippro/com/iprndMain/selectBusinessAnnounceList.do?bbsType=bs"
page = requests.get(url)
soup = bs(page.text, "html.parser")

# 제목 생성
a = soup.select('td.left a')
s1 = []
for tag in a :
    s1.append ( tag.text.replace('\n','').replace('\t','').replace('  ','').replace('\xa0','').replace('\r','') )

a = pd.DataFrame ( s1 )
# 번호 생성
b = soup.select('body div.main_contents div.sub_contents form table tbody tr td:nth-child(1)')
s1 = []
s1 = [int(td.get_text(strip=True)) for td in b]
b = pd.DataFrame ( s1 )
# 등록일 생성
c = soup.select ( 'body  div.main_contents  div.sub_contents  form  table  tbody  tr  td:nth-child(5)' )
start_dates = []
end_dates = []
s1=[]
for tag in c:
    s1.append(tag.text.replace("<td>", "").replace("</td>", "").replace("<br>", "").replace("</br>", ""))

for tag in s1 :
    start_date, end_date = tag.split(" ~ ")
    start_dates.append(start_date)
    end_dates.append(end_date)

c=pd.DataFrame(start_dates)
Deadline=pd.DataFrame(end_dates)

time_format = "%Y-%m-%d"
s1 = []
for tag in c[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

c = pd.DataFrame ( s1 )

s1 = []
for tag in Deadline[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

Deadline = pd.DataFrame ( s1 )

# 조회수생성
d = soup.select ( 'body  div.main_contents  div.sub_contents form  table  tbody  tr td:nth-child(4)' )
s1 = []
s1 = [int(td.get_text(strip=True)) for td in d]
d = pd.DataFrame ( s1 )

# Condition
e = soup.select ( 'body  div.main_contents  div.sub_contents  form  table  tbody  tr td.left  a  img' )
s1=[]
for tag in e :
    s1.append(tag.get('alt'))

Condition = pd.DataFrame ( s1 )

# 구분생성
e= soup.select ( 'body  div.main_contents  div.sub_contents  form  table  tbody  tr td:nth-child(2)' )
s1 = []
s1 = [td.get_text(strip=True) for td in e]
e = pd.DataFrame ( s1 )

#컬럼내임 변경
a.rename(columns={ 0: 'a'}, inplace=True)
b.rename(columns={ 0: 'b'}, inplace=True)
c.rename(columns={ 0: 'c'}, inplace=True)
Deadline.rename(columns={ 0: 'Deadline'}, inplace=True)
Condition.rename(columns={ 0: 'Condition'}, inplace=True)


df=[]
df= pd.merge(e,a, left_index=True, right_index=True, how='outer')
df= pd.merge(df,b, left_index=True, right_index=True, how='outer')
df= pd.merge(df,c, left_index=True, right_index=True, how='outer')
df['link']= url
df['Site']= '전략개발원_사업공고'
df= pd.merge(df,Deadline, left_index=True, right_index=True, how='outer')
df= pd.merge(df,Condition, left_index=True, right_index=True, how='outer')

df.columns=['Uploader','Headline', "Number", "Date1", "Link", "Site", 'Deadline', 'Condition']
전략개발원_사업공고=df[["Site","Number",'Headline',"Link",'Uploader',"Date1", 'Deadline', 'Condition' ]]


# < 전략개발원_사업공고 - -----------------------------------------------------------------------------------완료 >

# < 전략개발원_협력기관공고 - -----------------------------------------------------------------------------------시작 4>

url="https://biz.kista.re.kr/ippro/com/iprndMain/selectBusinessAnnounceList.do?bbsType=ac"
page = requests.get(url)
soup = bs(page.text, "html.parser")

# 제목 생성
a = soup.select('td.left a')
s1 = []
for tag in a :
    s1.append ( tag.text.replace('\n','').replace('\t','').replace('  ','').replace('\xa0','').replace('\r','') )

a = pd.DataFrame ( s1 )
# 번호 생성
b = soup.select('body div.main_contents div.sub_contents form table tbody tr td:nth-child(1)')
s1 = []
s1 = [int(td.get_text(strip=True)) for td in b]
b = pd.DataFrame ( s1 )
# 등록일 생성
c = soup.select ( 'body  div.main_contents  div.sub_contents  form  table  tbody  tr  td:nth-child(5)' )
start_dates = []
end_dates = []
s1=[]
for tag in c:
    s1.append(tag.text.replace("<td>", "").replace("</td>", "").replace("<br>", "").replace("</br>", ""))

for tag in s1 :
    start_date, end_date = tag.split(" ~ ")
    start_dates.append(start_date)
    end_dates.append(end_date)

c=pd.DataFrame(start_dates)
Deadline=pd.DataFrame(end_dates)

time_format = "%Y-%m-%d"
s1 = []
for tag in c[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

c = pd.DataFrame ( s1 )

s1 = []
for tag in Deadline[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

Deadline = pd.DataFrame ( s1 )

# 조회수생성
d = soup.select ( 'body  div.main_contents  div.sub_contents form  table  tbody  tr td:nth-child(4)' )
s1 = []
s1 = [int(td.get_text(strip=True)) for td in d]
d = pd.DataFrame ( s1 )

# Condition
e = soup.select ( 'body  div.main_contents  div.sub_contents  form  table  tbody  tr td.left  a  img' )
s1=[]
for tag in e :
    s1.append(tag.get('alt'))

Condition=pd.DataFrame(s1)

# 구분생성
e= soup.select ( 'body  div.main_contents  div.sub_contents  form  table  tbody  tr td:nth-child(2)' )
s1 = []
s1 = [td.get_text(strip=True) for td in e]
e = pd.DataFrame ( s1 )

#컬럼이름 변경
a.rename(columns={ 0: 'a'}, inplace=True)
b.rename(columns={ 0: 'b'}, inplace=True)
c.rename(columns={ 0: 'c'}, inplace=True)
Deadline.rename(columns={ 0: 'Deadline'}, inplace=True)
Condition.rename(columns={ 0: 'Condition'}, inplace=True)

df=[]
df= pd.merge(e,a, left_index=True, right_index=True, how='outer')
df= pd.merge(df,b, left_index=True, right_index=True, how='outer')
df= pd.merge(df,c, left_index=True, right_index=True, how='outer')
df['link']= url
df['Site']= '전략개발원_협력기관모집공고'
df= pd.merge(df,Deadline, left_index=True, right_index=True, how='outer')
df= pd.merge(df,Condition, left_index=True, right_index=True, how='outer')

df.columns=['Uploader','Headline', "Number", "Date1", "Link", "Site", 'Deadline', 'Condition']

전략개발원_협력기관모집공고=df[["Site","Number",'Headline',"Link",'Uploader',"Date1", 'Deadline', 'Condition' ]]


# < 전략개발원_협력기관공고 - -----------------------------------------------------------------------------------종료 >


# < 지식재산바우처_스타트업 --------------------------------------------------------------------------------------- 시작 5>

url="https://ripc.org/ipvoucher/notiStatupList.do"
page = requests.get(url)
soup = bs(page.text, "html.parser")


# 제목 생성
a = soup.select('body  div.center_form_box  form  div.startup_list  div div.startup_txt  p:nth-child(2)  a')
s1 = []
s1 = [td.get_text(strip=True) for td in a]
a = pd.DataFrame ( s1 )
# 번호 생성
b=len(a)
s1=list(range(1, b+1))
b = pd.DataFrame ( s1 )
# 등록일 생성
c = soup.select ( 'body  div.center_form_box  form  div.startup_list  div div.startup_txt  p:nth-child(3)' )
start_dates = []
end_dates = []
s1=[]
for tag in c:
    s1.append(tag.text.replace("<td>", "").replace("</td>", "").replace("<br>", "").replace("</br>", "").replace("접수기간", "").replace(":", "").replace(" ", ""))

for tag in s1 :
    start_date, end_date = tag.split("~")
    start_dates.append(start_date)
    end_dates.append(end_date)

c=pd.DataFrame(start_dates)
Deadline=pd.DataFrame(end_dates)

time_format = "%Y.%m.%d%H%M"
s1 = []
for tag in c[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

c = pd.DataFrame ( s1 )

s1 = []
for tag in Deadline[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

Deadline = pd.DataFrame ( s1 )

# 조회수생성
d = soup.select ( 'body  div.center_form_box  form  div.startup_list  div  div.startup_option  div.view_count' )
s1 = []
for tag in d:
    s1.append(int(tag.text.replace("조회수", "").replace(":", "").replace(" ", "")))

d = pd.DataFrame ( s1 )

# Condition
e = soup.select ( 'body  div.center_form_box  form  div.startup_list  div div.startup_icon.close' )
s1=[]
s1 = [td.get_text(strip=True) for td in e]
e = pd.DataFrame ( s1 )


#컬럼이름 변경
a.rename(columns={ 0: 'a'}, inplace=True)
b.rename(columns={ 0: 'b'}, inplace=True)
c.rename(columns={ 0: 'c'}, inplace=True)
Deadline.rename(columns={ 0: 'Deadline'}, inplace=True)
e.rename(columns={ 0: 'e'}, inplace=True)

# a: 제목 생성 b:번호 c:등록일 d: 조회수 Deadline:마감일 e:현재상태

df=[]

df= pd.merge(a,b, left_index=True, right_index=True, how='outer')
df['Uploader'] = ""
df= pd.merge(df,c, left_index=True, right_index=True, how='outer')
df['link']= url
df['Site']= '지식재산바우처_스타트업'
df= pd.merge(df,Deadline, left_index=True, right_index=True, how='outer')
df= pd.merge(df,e, left_index=True, right_index=True, how='outer')

df.columns=['Headline', "Number", 'Uploader', "Date1", "Link", "Site", 'Deadline', 'Condition']

지식재산바우처_스타트업=df[["Site", "Number", 'Headline', "Link", 'Uploader',"Date1", 'Deadline', 'Condition' ]]

# < 지식재산바우처_스타트업 ------------------------------------------------------------------------------------ 종료 >

# < 지식재산바우처_IP서비스기관모집 --------------------------------------------------------------------------------- 시작 6>

url="https://ripc.org/ipvoucher/ipSrvc/notiIpSrvcList.do"
page = requests.get(url)
soup = bs(page.text, "html.parser")


# 제목 생성
a = soup.select('#notiIpSrvcListForm  div div.startup_list  div div.startup_txt  p:nth-child(2)  a')
s1 = []
s1 = [td.get_text(strip=True) for td in a]
a = pd.DataFrame ( s1 )
# 번호 생성
b=len(a)
s1=list(range(1, b+1))
b = pd.DataFrame ( s1 )
# 등록일 생성
c = soup.select ( '#notiIpSrvcListForm  div  div.startup_list  div div.startup_txt  p:nth-child(3)' )
start_dates = []
end_dates = []
s1=[]
for tag in c:
    s1.append(tag.text.replace("접수기간", "").replace(":", "").replace(" ", "").replace("\n", "").replace("\r", "").replace("\t", "").replace("\xa0", ""))

for tag in s1 :
    start_date, end_date = tag.split("~")
    start_dates.append(start_date)
    end_dates.append(end_date)

c=pd.DataFrame(start_dates)
Deadline=pd.DataFrame(end_dates)

time_format = "%Y.%m.%d%H%M"
s1 = []
for tag in c[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

c = pd.DataFrame ( s1 )

s1 = []
for tag in Deadline[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

Deadline = pd.DataFrame ( s1 )

# 조회수생성
d = soup.select ( '#notiIpSrvcListForm  div  div.startup_list  div div.startup_option  div.view_count' )
s1 = []
for tag in d:
    s1.append(int(tag.text.replace("조회수", "").replace(":", "").replace(" ", "")))

d = pd.DataFrame ( s1 )

# Condition
e = soup.select ( '#notiIpSrvcListForm  div  div.startup_list  div div.startup_icon.close' )
s1=[]
s1 = [td.get_text(strip=True) for td in e]
e = pd.DataFrame ( s1 )


#컬럼이름 변경
a.rename(columns={ 0: 'a'}, inplace=True)
b.rename(columns={ 0: 'b'}, inplace=True)
c.rename(columns={ 0: 'c'}, inplace=True)
Deadline.rename(columns={ 0: 'Deadline'}, inplace=True)
e.rename(columns={ 0: 'e'}, inplace=True)

# a: 제목 생성 b:번호 c:등록일 d: 조회수 Deadline:마감일 e:현재상태

df=[]

df= pd.merge(a,b, left_index=True, right_index=True, how='outer')
df['Uploader'] = ""
df= pd.merge(df,c, left_index=True, right_index=True, how='outer')
df['link']= url
df['Site']= '지식재산바우처_서비스기관모집'
df= pd.merge(df,Deadline, left_index=True, right_index=True, how='outer')
df= pd.merge(df,e, left_index=True, right_index=True, how='outer')

df.columns=['Headline', "Number", 'Uploader', "Date1", "Link", "Site", 'Deadline', 'Condition']

지식재산바우처_IP서비스기관모집=df[["Site", "Number", 'Headline', "Link", 'Uploader',"Date1", 'Deadline', 'Condition' ]]




# < 지식재산바우처_IP서비스기관모집 --------------------------------------------------------------------------------- 종료 >

# < KAUTM ----------------------------------------------------------------------------------------------------- 시작 7>

url="http://www.kautm.net/bbs/?so_table=tlo_news&category=business"

page = requests.get(url)

soup = bs(page.content.decode('utf-8','replace'), "html.parser" )

#헤드라인
a=soup.select('body > div > section > section.sec-centent > div.page-conts-wrap > div > div.boardListArea > div.srboardList > div.listTable > table > tbody > tr > td.title > a > span.board-subject')
s1=[]
for tag in a :
    s1.append(tag.text)

a= pd.DataFrame(s1)

#링크
b=soup.select('body > div > section > section.sec-centent > div.page-conts-wrap > div > div.boardListArea > div.srboardList > div.listTable > table > tbody > tr > td.title > a')
s1=[]
for tag in b :
    s1.append('www.kautm.net'+tag["href"])

b= pd.DataFrame(s1)

# Uploader, 날짜, 조회수
c=soup.select('body > div > section > section.sec-centent > div.page-conts-wrap > div > div.boardListArea > div.srboardList > div.listTable > table > tbody > tr > td.mob-none:nth-child(3)')
s1=[]
s2=[]
s3=[]
s4=[]

for tag in c :
    s1.append(tag.text)

c= pd.DataFrame(s1)

v_split = c[0].str.split('\n')
#uploader
c = pd.DataFrame(v_split.str.get(0))
#조회수
e = pd.DataFrame(v_split.str.get(2))
#날짜
d = pd.DataFrame(v_split.str.get(1))
s1=[]
time_format= "%Y-%m-%d"
for tag in d[0]:
      s1.append ( datetime.datetime.strptime ( tag , time_format ) )

d = pd.DataFrame(s1)

# Number
f=soup.select('body > div > section > section.sec-centent > div.page-conts-wrap > div > div.boardListArea > div.srboardList > div.listTable > table > tbody > tr > td:nth-child(1)')
s1=[]
for tag in f :
    s1.append(tag.text)

f = pd.DataFrame(s1)


#컬럼이름 변경
a.rename(columns={ 0: 'a'}, inplace=True)
d.rename(columns={ 0: 'd'}, inplace=True)
b.rename(columns={ 0: 'b'}, inplace=True)
c.rename(columns={ 0: 'c'}, inplace=True)
f.rename(columns={ 0: 'f'}, inplace=True)

df= pd.merge(a,f, left_index=True, right_index=True, how='outer')
df= pd.merge(df,c, left_index=True, right_index=True, how='outer')
df= pd.merge(df,d, left_index=True, right_index=True, how='outer')
df= pd.merge(df,b, left_index=True, right_index=True, how='outer')
df['Site']= 'KAUTM'
df['Deadline']= ''
df['Condition']= ''

df.columns=['Headline', "Number", 'Uploader', "Date1", "Link", "Site", 'Deadline', 'Condition']

KAUTM=df[["Site", "Number", 'Headline', "Link", 'Uploader',"Date1", 'Deadline', 'Condition' ]]



# < KAUTM --------------------------------------------------------------------------------- 종료 7>
# < 수출바우처 공지 - -----------------------------------------------------------------------------------시작 8>

chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=chrome_options)

driver.get('https://www.exportvoucher.com/portal/board/boardList?bbs_id=1&active_menu_cd=EZ005004000')
time.sleep(1)


def function_exportvoucher() :
    page = driver.page_source
    soup = bs ( page , "html.parser" )
    # 제목 생성
    a =  soup.select('body>div>section>div>article>form>table>tbody>tr>td>a')
    s1=[]
    for tag in a :
        s1.append(tag.text)
    a= pd.DataFrame(s1)
    # 번호 생성
    b =  soup.select('body>div>section>div>article>form>table>tbody>tr>td:nth-child(1)')
    s1=[]
    for tag in b :
        s1.append(tag.text)
    b= pd.DataFrame(s1)
    # 등록일 생성
    c =  soup.select('body>div>section>div>article>form>table>tbody>tr>td:nth-child(3)')
    s1=[]
    for tag in c :
        s1.append(tag.text)
    c= pd.DataFrame(s1)
    time_format = "%Y-%m-%d"
    s1=[]
    for tag in c[0]:
        s1.append ( datetime.datetime.strptime ( tag , time_format ) )
    c=pd.DataFrame(s1)
    #조회수생성
    d =  soup.select('body>div>section>div>article>form>table>tbody>tr>td:nth-child(4)')
    s1=[]
    for tag in d :
        s1.append(tag.text)
    d = pd.DataFrame(s1)
    # 컬럼이름변경
    a.rename(columns={0: 'a'}, inplace=True)
    b.rename(columns={0: 'b'}, inplace=True)
    c.rename(columns={0: 'c'}, inplace=True)
    d.rename(columns={0: 'd'}, inplace=True)
    df= pd.merge(a,b, left_index=True, right_index=True, how='outer')
    df['Uploader']= ''
    df= pd.merge(df,c, left_index=True, right_index=True, how='outer')
    df['Link']= 'https://www.exportvoucher.com/portal/board/boardList'
    df['Site']= '수출바우처_공지'
    df['Deadline']= ''
    df= pd.merge(df,d, left_index=True, right_index=True, how='outer')
    df.columns=['Headline', "Number", 'Uploader', "Date1", "Link", "Site", 'Deadline', 'Condition']
    df=df[["Site", "Number", 'Headline', "Link", 'Uploader',"Date1", 'Deadline', 'Condition' ]]
    return df


# 1번
Export_voucher=[]
Export_voucher = function_exportvoucher() # 1번

# 2번
driver.find_element(By.XPATH,'//*[@id="contents"]/form/div[3]/div/a[2]').click()
driver.implicitly_wait(10)
a = function_exportvoucher()
Export_voucher=pd.concat([Export_voucher,a])

# 3번

driver.find_element(By.XPATH,'//*[@id="contents"]/form/div[3]/div/a[4]').click()
driver.implicitly_wait(10)
a = function_exportvoucher()
Export_voucher=pd.concat([Export_voucher,a])

# 4번

driver.find_element(By.XPATH,'//*[@id="contents"]/form/div[3]/div/a[5]').click()
driver.implicitly_wait(10)
a = function_exportvoucher()
Export_voucher=pd.concat([Export_voucher,a])

# 5번

driver.find_element(By.XPATH,'//*[@id="contents"]/form/div[3]/div/a[6]').click()
driver.implicitly_wait(10)
a = function_exportvoucher()
Export_voucher=pd.concat([Export_voucher,a])

# 6번

driver.find_element(By.XPATH,'//*[@id="contents"]/form/div[3]/div/a[7]').click()
driver.implicitly_wait(10)
a = function_exportvoucher()
Export_voucher=pd.concat([Export_voucher,a])

# 7번

driver.find_element(By.XPATH,'//*[@id="contents"]/form/div[3]/div/a[8]').click()
driver.implicitly_wait(10)
a = function_exportvoucher()
Export_voucher=pd.concat([Export_voucher,a])

# 8번

driver.find_element(By.XPATH,'//*[@id="contents"]/form/div[3]/div/a[9]').click()
driver.implicitly_wait(10)
a = function_exportvoucher()
Export_voucher=pd.concat([Export_voucher,a])

# 9번

driver.find_element(By.XPATH,'//*[@id="contents"]/form/div[3]/div/a[10]').click()
driver.implicitly_wait(10)
a = function_exportvoucher()
Export_voucher=pd.concat([Export_voucher,a])

# 10번

driver.find_element(By.XPATH,'//*[@id="contents"]/form/div[3]/div/a[11]').click()
driver.implicitly_wait(10)
a = function_exportvoucher()
Export_voucher=pd.concat([Export_voucher,a])

s1=Export_voucher.drop_duplicates(['Headline', 'Date1'])
s1= s1.reset_index(drop=True)

#조회수 숫자로 변경
s2=[]
for tag in s1.Condition:
    s2.append(int(tag))

s2=pd.DataFrame(s2)
s2.columns=['Condition']
s1.Condition=s2.Condition


#중요 삽입
s2=s1.Number
s2=pd.DataFrame(s2)

for tag in range(len(s2)) :
    if s2['Number'][tag].isdecimal() == False :
        s2['Number'][tag] = '중요'

s2.columns=['Number']
s1.Number=s2.Number

Export_voucher = s1

# < 수출바우처 공지 - -----------------------------------------------------------------------------------종료 8>

# < 저장 ------------------------------------------------------------------------------------------------------ 저장 >

ct=datetime.datetime.now()
ctstr=ct.strftime("%Y년 %m월 %d일 %H시%M분%S초")

End_result=[]
# End_result=pd.concat([IP_NAVI_사업공고,IP_NAVI_입찰공고])
# End_result=pd.concat([End_result,RIPC_사업공고])
End_result=pd.concat([RIPC_사업공고,IP_NAVI_사업공고])
End_result=pd.concat([End_result,전략개발원_사업공고])
End_result=pd.concat([End_result,전략개발원_협력기관모집공고])
# End_result=pd.concat([End_result,지식재산바우처_스타트업])
# End_result=pd.concat([End_result,지식재산바우처_IP서비스기관모집])
End_result=pd.concat([End_result,KAUTM])
End_result=pd.concat([End_result,Export_voucher])

#시간순으로 정렬
전략개발원_협력기관모집공고=전략개발원_협력기관모집공고.sort_values('Date1',ascending=False)
전략개발원_사업공고=전략개발원_사업공고.sort_values('Date1',ascending=False)
RIPC_사업공고=RIPC_사업공고.sort_values('Date1',ascending=False)
IP_NAVI_사업공고=IP_NAVI_사업공고.sort_values('Date1',ascending=False)
# IP_NAVI_입찰공고=IP_NAVI_입찰공고.sort_values('Date1',ascending=False)
지식재산바우처_스타트업=지식재산바우처_스타트업.sort_values('Date1',ascending=False)
지식재산바우처_IP서비스기관모집=지식재산바우처_IP서비스기관모집.sort_values('Date1',ascending=False)
End_result=End_result.sort_values('Date1',ascending=False)
KAUTM=KAUTM.sort_values('Date1',ascending=False)
Export_voucher=Export_voucher.sort_values('Date1',ascending=False)

#중복제거
End_result=End_result.drop_duplicates(["Headline","Uploader"], keep='first',  ignore_index=True)
RIPC_사업공고=RIPC_사업공고.drop_duplicates(["Headline","Uploader"], keep='first',  ignore_index=True)

writer=pd.ExcelWriter('C:/작업서류SG/작업서류/0.Crawling/정부사업크롤링('+ctstr+').xlsx', engine='openpyxl')

End_result.to_excel(writer, sheet_name='전체')
전략개발원_협력기관모집공고.to_excel(writer, sheet_name='전략개발원_협력기관모집공고')
전략개발원_사업공고.to_excel(writer, sheet_name='전략개발원_사업공고')
RIPC_사업공고.to_excel(writer, sheet_name='RIPC_사업공고')
IP_NAVI_사업공고.to_excel(writer, sheet_name='IP_NAVI_사업공고')
# IP_NAVI_입찰공고.to_excel(writer, sheet_name='IP_NAVI_입찰공고')
지식재산바우처_스타트업.to_excel(writer, sheet_name='지식재산바우처_스타트업')
지식재산바우처_IP서비스기관모집.to_excel(writer, sheet_name='지식재산바우처_IP서비스기관모집')
KAUTM.to_excel(writer, sheet_name='KAUTM')
Export_voucher.to_excel(writer, sheet_name='Export_voucher')
writer.close()

# 엑셀 간격 조정
from openpyxl import load_workbook
wb = load_workbook('C:/작업서류SG/작업서류/0.Crawling/정부사업크롤링('+ctstr+').xlsx')
for tag in wb.sheetnames:
    ws = wb[tag]
    ws.column_dimensions['A'].width=3
    ws.column_dimensions['B'].width=20
    ws.column_dimensions['C'].width=20
    ws.column_dimensions['D'].width=80
    ws.column_dimensions['E'].width=20
    ws.column_dimensions['F'].width=10
    ws.column_dimensions['G'].width=20
    ws.column_dimensions['H'].width=20
    ws.column_dimensions['I'].width=10

wb.save('C:/작업서류SG/작업서류/0.Crawling/정부사업크롤링('+ctstr+').xlsx')


