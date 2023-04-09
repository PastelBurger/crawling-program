

import requests
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup as bs
from pandas import DataFrame
from selenium import webdriver
import time
import openpyxl
import datetime
from datetime import timedelta


# < IP실시간뉴스 --------------------------------------------------------------------------------- 시작 1>


options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
driver = webdriver.Chrome(options=options)
driver.implicitly_wait(3)
driver.get('http://www.ip-navi.or.kr/ipnavi/ref/boardList.navi?boardCode=B00017')
time.sleep(1)

def function_IP_news() :
    page = driver.page_source
    soup = bs(page, "html.parser")
    #헤드라인
    a = soup.select('body > div.container > div.board_wrap > table > tbody > tr> td.align_L > a')
    s1=[]
    for tag in a :
        s1.append(tag.text)
    s2=[]
    for tag in s1 :
        s2.append(tag.replace('\n','').replace('\t','').replace('  ','').replace('\xa0','').replace('\r',''))

    a= pd.DataFrame(s2)

    #링크
    b = soup.select('body > div.container > div.board_wrap > table > tbody > tr> td.align_L > a')
    s1=[]
    for tag in b :
        s1.append(tag["href"])

    s1=pd.DataFrame(s1)
    v_split = s1[0].str.split('(')
    s1= pd.DataFrame(v_split.str.get(1))

    v_split = s1[0].str.split(',')
    s1= pd.DataFrame(v_split.str.get(0))
    s2=[]
    for tag in s1[0] :
        s2.append(tag[1:-1])

    b= pd.DataFrame(s2)

    # Uploader
    s1=[]
    for tag in range(len(a)) :
        s1.append("IP_NAVI_NEWS")

    c= pd.DataFrame(s1)

    # 날짜
    d=soup.select('body > div.container > div.board_wrap > table > tbody > tr > td:nth-child(3)')
    d= pd.DataFrame(d)
    s1=[]
    time_format= "%Y-%m-%d"
    for tag in d[0]:
          s1.append ( datetime.datetime.strptime ( tag , time_format ) )

    d= pd.DataFrame(s1)

    # 조회수
    s1=[]
    for tag in range(len(a)) :
        s1.append("None")

    e= pd.DataFrame(s1)


    # 분류
    s1=[]
    for tag in range(len(a)) :
        s1.append("None")

    f= pd.DataFrame(s1)

    #번호
    g=soup.select('body > div.container > div.board_wrap > table > tbody > tr > td:nth-child(1)')
    g= pd.DataFrame(g)

    #헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
    df=[]
    df= pd.merge(g,f, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,a, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,b, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,c, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,e, left_index=True, right_index=True, how='outer')
    df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )
    return df


from selenium.webdriver.common.keys import Keys

# 1번
IP_news=[]
IP_news = function_IP_news() # 1번

# 2번
driver.find_element_by_xpath('//*[@id="divPagination"]/a[1]').send_keys(Keys.ENTER)
driver.implicitly_wait(3)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 3번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[4]').send_keys(Keys.ENTER)
driver.implicitly_wait(3)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 4번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[5]').send_keys(Keys.ENTER)
driver.implicitly_wait(3)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 5번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[6]').send_keys(Keys.ENTER)
driver.implicitly_wait(3)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 6번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[7]').send_keys(Keys.ENTER)
driver.implicitly_wait(3)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 7번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[8]').send_keys(Keys.ENTER)
driver.implicitly_wait(3)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 8번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[9]').send_keys(Keys.ENTER)
driver.implicitly_wait(3)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 9번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[9]').send_keys(Keys.ENTER)
driver.implicitly_wait(3)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 10번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[9]').send_keys(Keys.ENTER)
driver.implicitly_wait(3)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

IP_news.columns=['번호', '분류', "제목", '링크',"작성자", "조회", "등록일"]
# IP_news.to_excel('./Text1.xlsx',sheet_name='전체')

# < IP_Desk_news --------------------------------------------------------------------------------- 종료 2>


options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
driver = webdriver.Chrome(options=options)
driver.implicitly_wait(3)
driver.get('http://www.ip-navi.or.kr/ipnavi/ref/boardList.navi?boardCode=B00018')
time.sleep(1)


def function_IP_Desk_news() :
    page = driver.page_source
    soup = bs ( page , "html.parser" )

    # 헤드라인
    a = []
    a = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr > td.align_L>a' )
    s1 = []
    for tag in a :
        s1.append ( tag.text )

    s2 = []
    for tag in s1 :
        s2.append (
            tag.replace ( '\n' , '' ).replace ( '\t' , '' ).replace ( '  ' , '' ).replace ( '\xa0' , '' ).replace (
                '\r' , '' ) )

    a = pd.DataFrame ( s2 )

    # 링크
    b = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr > td.align_L >a ' )
    s1 = []
    for tag in b :
        s1.append ( tag["href"] )

    s1 = pd.DataFrame ( s1 )
    v_split = s1[0].str.split ( ',' )
    s1 = pd.DataFrame ( v_split.str.get ( 1 ) )
    v_split = s1[0].str.split ( ')' )
    s1 = pd.DataFrame ( v_split.str.get ( 0 ) )
    s2 = []
    for tag in s1[0] :
        s2.append ( tag[1 :-1] )
    s1 = []
    for tag in s2 :
        s1.append (
            "https://www.ip-navi.or.kr/ipnavi/ref/boardDetail.navi;jsessionid=2D70B0DCB138F2CF8F99F936475D17F6?boardCode=B00018&boardSeq=" + tag )

    b = pd.DataFrame ( s1 )

    # Uploader
    s1 = []
    for tag in range ( len ( a ) ) :
        s1.append ( "IP_DESK_NEWS" )
    c = pd.DataFrame ( s1 )
    # 날짜
    d = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr > td:nth-child(4)' )
    d = pd.DataFrame ( d )
    s1 = []
    time_format = "%Y-%m-%d"
    for tag in d[0] :
        s1.append ( datetime.datetime.strptime ( tag , time_format ) )

    d = pd.DataFrame ( s1 )
    # 조회수
    e = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr > td:nth-child(3)' )
    e = pd.DataFrame ( e )
    s1 = []
    for tag in e[0] :
        s1.append ( int ( tag ) )

    e = pd.DataFrame ( s1 )
    s1 = []
    for tag in e[0] :
        s1.append ( int ( tag ) )

    e = pd.DataFrame ( s1 )

    # 분류
    s1 = []
    for tag in range ( len ( a ) ) :
        s1.append ( "NONE" )
    f = pd.DataFrame ( s1 )
    # 번호
    g = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr > td:nth-child(1)' )
    g = pd.DataFrame ( g )

    # 헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
    df = []
    df = g
    df['분류'] = 'IP_DESK_지재권뉴스'
    df = pd.merge ( df , a , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , b , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , c , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , e , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )
    return df


#1
IP_Desk_news=[]
IP_Desk_news = function_IP_Desk_news() # 1번

# 2번
driver.find_element_by_xpath('//*[@id="divPagination"]/a[1]').send_keys(Keys.ENTER)
driver.implicitly_wait(3)
a = function_IP_Desk_news()
IP_Desk_news=pd.concat([IP_Desk_news,a])

# 3번
driver.find_element_by_xpath('//*[@id="divPagination"]/a[4]').send_keys(Keys.ENTER)
driver.implicitly_wait(3)
a = function_IP_Desk_news()
IP_Desk_news=pd.concat([IP_Desk_news,a])

# 4번
driver.find_element_by_xpath('//*[@id="divPagination"]/a[5]').send_keys(Keys.ENTER)
driver.implicitly_wait(3)
a = function_IP_Desk_news()
IP_Desk_news=pd.concat([IP_Desk_news,a])

# 5번
driver.find_element_by_xpath('//*[@id="divPagination"]/a[6]').send_keys(Keys.ENTER)
driver.implicitly_wait(3)
a = function_IP_Desk_news()
IP_Desk_news=pd.concat([IP_Desk_news,a])


IP_Desk_news.columns=['번호', '분류', "제목", '링크',"작성자", "조회", "등록일"]
# IP_Desk_news.to_excel('./Text2.xlsx',sheet_name='전체')

# < IP_Desk_news --------------------------------------------------------------------------------- 종료 2>
# < 지재권 판례 --------------------------------------------------------------------------------- 시작 3>
# < 지재권 판례 1 >
driver.get('https://www.ip-navi.or.kr/ipnavi/precedent/insightList.navi?boardCode=B00006')
driver.implicitly_wait(3)
#
# url="https://www.ip-navi.or.kr/board/boardList.navi?boardCode=B00006"
# page = requests.get(url)
# soup = bs(page.text, "html.parser")


# def function_IP_judicial_precedent(ss) :
page = driver.page_source
soup = bs ( page , "html.parser" )

# 헤드라인
a = []
a = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr > td.subject.align_L > a' )
s1 = []
for tag in a :
    s1.append ( tag.text )

s2 = []
for tag in s1 :
    s2.append (
        tag.replace ( '\n' , '' ).replace ( '\t' , '' ).replace ( '  ' , '' ).replace ( '\xa0' , '' ).replace (
            '\r' , '' ) )

a = pd.DataFrame ( s2 )

# 링크
b = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr > td.subject.align_L > a' )
s1 = []
for tag in b :
    s1.append ( tag["href"] )

s1 = pd.DataFrame ( s1 )
v_split = s1[0].str.split ( ',' )
s1 = pd.DataFrame ( v_split.str.get ( 1 ) )
v_split = s1[0].str.split ( ')' )
s1 = pd.DataFrame ( v_split.str.get ( 0 ) )
s2 = []
for tag in s1[0] :
    s2.append ( tag[1 :-1] )

# https://www.ip-navi.or.kr/ipnavi/ref/boardDetail.navi;jsessionid=2D70B0DCB138F2CF8F99F936475D17F6?boardCode=B00018&boardSeq=" + tag

s1 = []
ss = str ( ss )
for tag in s2 :
    s1.append ("https://www.ip-navi.or.kr/board/viewBoardArticle.navi?boardCode=B00006&getMaxInactiveinterval=0&boardSeq=" + tag + "&pageNum=" + ss + "&national=&searchCodeArrayValue1=TOTAL1&searchCodeArrayValue2=TOTAL2&searchCodeArrayValue3=TOTAL3&searchCodeArrayValue4=TOTAL4&search_keyword=&searchCode1=TOTAL1&searchCode2=TOTAL2&legal_issue=&searchCode3=TOTAL3&searchCode4=TOTAL4&judgment_start_date=&judgment_end_date=&register_start_date=&register_end_date=" )

b = pd.DataFrame ( s1 )

# 업로더
 # c

# 날짜
d = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(5)' )
d = pd.DataFrame ( d )
s1 = []
time_format = "%Y-%m-%d"
for tag in d[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

d = pd.DataFrame ( s1 )

# 조회수
e = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(4)' )
e = pd.DataFrame ( e )
s1 = []
for tag in e[0] :
    s1.append ( int ( tag ) )

e = pd.DataFrame ( s1 )
s1 = []
for tag in e[0] :
    s1.append ( int ( tag ) )

e = pd.DataFrame ( s1 )

# 분류
f = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(2)' )
f = pd.DataFrame ( f )

# 번호
g = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(1)' )
g = pd.DataFrame ( g )

# 헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
df = []
df = g
df = pd.merge ( df , f , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , a , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , b , left_index=True , right_index=True , how='outer' )
df['업로더'] = '지재권판례'
df = pd.merge ( df , e , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )
    # return df

#1

ss=1
IP_jp = function_IP_judicial_precedent(ss) # 1번

# 2번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[2]').click()
driver.implicitly_wait(3)
ss=2
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 3번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[4]').click()
driver.implicitly_wait(3)
ss=3
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 4번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[5]').click()
driver.implicitly_wait(3)
ss=4
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 5번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[6]').click()
driver.implicitly_wait(3)
ss=5
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 6번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[7]').click()
driver.implicitly_wait(3)
ss=6
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 7번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[8]').click()
driver.implicitly_wait(3)
ss=7
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 8번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[9]').click()
driver.implicitly_wait(3)
ss=8
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 9번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[10]').click()
driver.implicitly_wait(3)
ss=9
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 10번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[11]').click()
driver.implicitly_wait(3)
ss=10
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

IP_jp.columns=['번호', '분류', "제목", '링크',"작성자", "조회", "등록일"]
# IP_jp.to_excel('./Text3.xlsx',sheet_name='전체')

# < 지재권 판례 2 - 실무적관점 >

driver.get('https://www.ip-navi.or.kr/board/boardList.navi?boardCode=B00030')
driver.implicitly_wait(3)

# url="https://www.ip-navi.or.kr/board/boardList.navi?boardCode=B00030"
# page = requests.get(url)
# soup = bs(page.text, "html.parser")

page = driver.page_source
soup = bs ( page , "html.parser" )

# 헤드라인
a = []
a=soup.select( '#contents > div.content_area > table > tbody > tr > td.td_tit' )
s1 = []
for tag in a :
    s1.append ( tag.text )

s2 = []
for tag in s1 :
    s2.append (
        tag.replace ( '\n' , '' ).replace ( '\t' , '' ).replace ( '  ' , '' ).replace ( '\xa0' , '' ).replace (
            '\r' , '' ) )

a = pd.DataFrame ( s2 )

# 링크
b = soup.select ( '#contents > div.content_area > table > tbody > tr > td.td_tit > a' )
s1 = []
for tag in b :
    s1.append ( tag["onclick"] )

s1 = pd.DataFrame ( s1 )
v_split = s1[0].str.split ( '(' )
s1 = pd.DataFrame ( v_split.str.get ( 1 ) )
v_split = s1[0].str.split ( ')' )
s1 = pd.DataFrame ( v_split.str.get ( 0 ) )
s2 = []
for tag in s1[0] :
    s2.append ( tag[1 :-1] )

s1 = []
for tag in s2 :
    s1.append (
        "https://www.ip-navi.or.kr/board/viewBoardArticle.navi?boardCode=B00030&getMaxInactiveinterval=0&boardSeq=" + tag + "&pageNum=1&national=&searchCodeArrayValue1=&searchCodeArrayValue2=&searchCodeArrayValue3=&searchCodeArrayValue4=&search_keyword=" )

b = pd.DataFrame ( s1 )

# 업로더
# c

# 날짜
d = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(5)' )
d = pd.DataFrame ( d )
s1 = []
time_format = "%Y-%m-%d"
for tag in d[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

d = pd.DataFrame ( s1 )

# 조회수
e = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(4)' )
e = pd.DataFrame ( e )
s1 = []
for tag in e[0] :
    s1.append ( int ( tag ) )

e = pd.DataFrame ( s1 )
s1 = []
for tag in e[0] :
    s1.append ( int ( tag ) )

e = pd.DataFrame ( s1 )

# 분류
 #f

# 번호
g = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(1)' )
g = pd.DataFrame ( g )

# 헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
df = []
df = g
df['분류'] = '실무적관점'
df = pd.merge ( df , a , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , b , left_index=True , right_index=True , how='outer' )
df['업로더'] = '실무적관점'
df = pd.merge ( df , e , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )

df.columns=['번호', '분류', "제목", '링크',"작성자", "조회", "등록일"]
IP_jp=pd.concat([IP_jp,df])


# < 지재권 판례 3 - 카드뉴스동영상 >

driver.get('https://www.ip-navi.or.kr/board/boardList.navi?boardCode=B00028')
driver.implicitly_wait(3)

# url="https://www.ip-navi.or.kr/board/boardList.navi?boardCode=B00028"
# page = requests.get(url)
# soup = bs(page.text, "html.parser")

page = driver.page_source
soup = bs ( page , "html.parser" )

# 헤드라인
a = []
a=soup.select( '#contents > div.content_area > table > tbody > tr > td.td_tit' )
s1 = []
for tag in a :
    s1.append ( tag.text )

s2 = []
for tag in s1 :
    s2.append (
        tag.replace ( '\n' , '' ).replace ( '\t' , '' ).replace ( '  ' , '' ).replace ( '\xa0' , '' ).replace (
            '\r' , '' ) )

a = pd.DataFrame ( s2 )

# 링크
b = soup.select ( '#contents > div.content_area > table > tbody > tr > td.td_tit > a' )
s1 = []
for tag in b :
    s1.append ( tag["onclick"] )

s1 = pd.DataFrame ( s1 )
v_split = s1[0].str.split ( '(' )
s1 = pd.DataFrame ( v_split.str.get ( 1 ) )
v_split = s1[0].str.split ( ')' )
s1 = pd.DataFrame ( v_split.str.get ( 0 ) )
s2 = []
for tag in s1[0] :
    s2.append ( tag[1 :-1] )

s1 = []
ss=1
ss = str ( ss )
for tag in s2 :
    s1.append (
        "https://www.ip-navi.or.kr/board/viewBoardArticle.navi?boardCode=B00028&getMaxInactiveinterval=0&boardSeq=" + tag + "&pageNum=" +ss+"&national=&searchCodeArrayValue1=&searchCodeArrayValue2=&searchCodeArrayValue3=&searchCodeArrayValue4=&search_keyword=" )

b = pd.DataFrame ( s1 )

# 업로더
# c

# 날짜
d = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(5)' )
d = pd.DataFrame ( d )
s1 = []
time_format = "%Y-%m-%d"
for tag in d[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

d = pd.DataFrame ( s1 )

# 조회수
e = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(4)' )
e = pd.DataFrame ( e )
s1 = []
for tag in e[0] :
    s1.append ( int ( tag ) )

e = pd.DataFrame ( s1 )
s1 = []
for tag in e[0] :
    s1.append ( int ( tag ) )

e = pd.DataFrame ( s1 )

# 분류
 #f

# 번호
g = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(1)' )
g = pd.DataFrame ( g )

# 헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
df = []
df = g
df['분류'] = '카드뉴스동영상'
df = pd.merge ( df , a , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , b , left_index=True , right_index=True , how='outer' )
df['업로더'] = '카드뉴스동영상'
df = pd.merge ( df , e , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )

df.columns=['번호', '분류', "제목", '링크',"작성자", "조회", "등록일"]
IP_jp=pd.concat([IP_jp,df])

## 카드뉴스 다음장
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[2]').click()
driver.implicitly_wait(3)

page = driver.page_source
soup = bs ( page , "html.parser" )

# 헤드라인
a = []
a=soup.select( '#contents > div.content_area > table > tbody > tr > td.td_tit' )
s1 = []
for tag in a :
    s1.append ( tag.text )

s2 = []
for tag in s1 :
    s2.append (
        tag.replace ( '\n' , '' ).replace ( '\t' , '' ).replace ( '  ' , '' ).replace ( '\xa0' , '' ).replace (
            '\r' , '' ) )

a = pd.DataFrame ( s2 )

# 링크
b = soup.select ( '#contents > div.content_area > table > tbody > tr > td.td_tit > a' )
s1 = []
for tag in b :
    s1.append ( tag["onclick"] )

s1 = pd.DataFrame ( s1 )
v_split = s1[0].str.split ( '(' )
s1 = pd.DataFrame ( v_split.str.get ( 1 ) )
v_split = s1[0].str.split ( ')' )
s1 = pd.DataFrame ( v_split.str.get ( 0 ) )
s2 = []
for tag in s1[0] :
    s2.append ( tag[1 :-1] )

s1 = []
ss=1
ss = str ( ss )
for tag in s2 :
    s1.append (
        "https://www.ip-navi.or.kr/board/viewBoardArticle.navi?boardCode=B00028&getMaxInactiveinterval=0&boardSeq=" + tag + "&pageNum=" +ss+"&national=&searchCodeArrayValue1=&searchCodeArrayValue2=&searchCodeArrayValue3=&searchCodeArrayValue4=&search_keyword=" )

b = pd.DataFrame ( s1 )

# 업로더
# c

# 날짜
d = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(5)' )
d = pd.DataFrame ( d )
s1 = []
time_format = "%Y-%m-%d"
for tag in d[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

d = pd.DataFrame ( s1 )

# 조회수
e = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(4)' )
e = pd.DataFrame ( e )
s1 = []
for tag in e[0] :
    s1.append ( int ( tag ) )

e = pd.DataFrame ( s1 )
s1 = []
for tag in e[0] :
    s1.append ( int ( tag ) )

e = pd.DataFrame ( s1 )

# 분류
 #f

# 번호
g = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(1)' )
g = pd.DataFrame ( g )

# 헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
df = []
df = g
df['분류'] = '카드뉴스동영상'
df = pd.merge ( df , a , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , b , left_index=True , right_index=True , how='outer' )
df['업로더'] = '카드뉴스동영상'
df = pd.merge ( df , e , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )

df.columns=['번호', '분류', "제목", '링크',"작성자", "조회", "등록일"]
IP_jp=pd.concat([IP_jp,df])

# IP_jp.to_excel('./Text3.xlsx',sheet_name='전체')



# < IP_Desk_news --------------------------------------------------------------------------------- 종료 2>
def function_IP_Desk_news(ss) :
    page = driver.page_source
    soup = bs ( page , "html.parser" )

    # 헤드라인
    a = []
    a = soup.select ( '#contents > div.content_area > table > tbody > tr > td.td_tit' )
    s1 = []
    for tag in a :
        s1.append ( tag.text )

    s2 = []
    for tag in s1 :
        s2.append (
            tag.replace ( '\n' , '' ).replace ( '\t' , '' ).replace ( '  ' , '' ).replace ( '\xa0' , '' ).replace (
                '\r' , '' ) )

    a = pd.DataFrame ( s2 )

    # 링크
    b = soup.select ( '#contents > div.content_area > table > tbody > tr > td.td_tit > a' )
    s1 = []
    for tag in b :
        s1.append ( tag["onclick"] )

    s1 = pd.DataFrame ( s1 )
    v_split = s1[0].str.split ( '(' )
    s1 = pd.DataFrame ( v_split.str.get ( 1 ) )
    v_split = s1[0].str.split ( ')' )
    s1 = pd.DataFrame ( v_split.str.get ( 0 ) )
    s2 = []
    for tag in s1[0] :
        s2.append ( tag[1 :-1] )
    s1 = []
    ss = str ( ss )
    for tag in s2 :
        s1.append (
            "https://www.ip-navi.or.kr/board/viewBoardArticle.navi?boardCode=B00018&getMaxInactiveinterval=0&boardSeq=" + tag + "&pageNum=" + ss + "&national=&searchCodeArrayValue1=&searchCodeArrayValue2=&searchCodeArrayValue3=&searchCodeArrayValue4=&search_keyword=&SelectedBox=B00018" )

    b = pd.DataFrame ( s1 )

    # 업로더
    c = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(3)' )
    c = pd.DataFrame ( c )
    # 날짜
    d = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(5)' )
    d = pd.DataFrame ( d )
    s1 = []
    time_format = "%Y-%m-%d"
    for tag in d[0] :
        s1.append ( datetime.datetime.strptime ( tag , time_format ) )

    d = pd.DataFrame ( s1 )
    # 조회수
    e = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(4)' )
    e = pd.DataFrame ( e )
    s1 = []
    for tag in e[0] :
        s1.append ( int ( tag ) )

    e = pd.DataFrame ( s1 )
    s1 = []
    for tag in e[0] :
        s1.append ( int ( tag ) )

    e = pd.DataFrame ( s1 )

    # 분류
    # f
    # 번호
    g = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(1)' )
    g = pd.DataFrame ( g )

    # 헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
    df = []
    df = g
    df['분류'] = 'IP_DESK_지재권뉴스'
    df = pd.merge ( df , a , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , b , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , c , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , e , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )
    return df


#1
driver.find_element_by_xpath('//*[@id="snb"]/ul/li[5]/ul/li[1]/a').click()
driver.implicitly_wait(3)
IP_Desk_news=[]
ss=1
IP_Desk_news = function_IP_Desk_news(ss) # 1번

# 2번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[2]').click()
driver.implicitly_wait(3)
ss=2
a = function_IP_Desk_news(ss)
IP_Desk_news=pd.concat([IP_Desk_news,a])

# 3번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[4]').click()
driver.implicitly_wait(3)
ss=3
a = function_IP_Desk_news(ss)
IP_Desk_news=pd.concat([IP_Desk_news,a])

# 4번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[5]').click()
driver.implicitly_wait(3)
ss=4
a = function_IP_Desk_news(ss)
IP_Desk_news=pd.concat([IP_Desk_news,a])

# 5번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[6]').click()
driver.implicitly_wait(3)
ss=5
a = function_IP_Desk_news(ss)
IP_Desk_news=pd.concat([IP_Desk_news,a])


IP_Desk_news.columns=['번호', '분류', "제목", '링크',"작성자", "조회", "등록일"]
# IP_Desk_news.to_excel('./Text2.xlsx',sheet_name='전체')

# < IP_Desk_news --------------------------------------------------------------------------------- 종료 2>
# < 지재권 판례 --------------------------------------------------------------------------------- 시작 3>
# < 지재권 판례 1 >
driver.get('https://www.ip-navi.or.kr/board/boardList.navi?boardCode=B00006')
driver.implicitly_wait(3)
#
# url="https://www.ip-navi.or.kr/board/boardList.navi?boardCode=B00006"
# page = requests.get(url)
# soup = bs(page.text, "html.parser")


def function_IP_judicial_precedent(ss) :
    page = driver.page_source
    soup = bs ( page , "html.parser" )

    # 헤드라인
    a = []
    a = soup.select ( '#contents > div.content_area > table > tbody > tr > td.td_tit' )
    s1 = []
    for tag in a :
        s1.append ( tag.text )

    s2 = []
    for tag in s1 :
        s2.append (
            tag.replace ( '\n' , '' ).replace ( '\t' , '' ).replace ( '  ' , '' ).replace ( '\xa0' , '' ).replace (
                '\r' , '' ) )

    a = pd.DataFrame ( s2 )

    # 링크
    b = soup.select ( '#contents > div.content_area > table > tbody > tr > td.td_tit > a' )
    s1 = []
    for tag in b :
        s1.append ( tag["href"] )

    s1 = pd.DataFrame ( s1 )
    v_split = s1[0].str.split ( '(' )
    s1 = pd.DataFrame ( v_split.str.get ( 1 ) )
    v_split = s1[0].str.split ( ')' )
    s1 = pd.DataFrame ( v_split.str.get ( 0 ) )
    s2 = []
    for tag in s1[0] :
        s2.append ( tag[1 :-1] )

    s1 = []
    ss = str ( ss )
    for tag in s2 :
        s1.append ("https://www.ip-navi.or.kr/board/viewBoardArticle.navi?boardCode=B00006&getMaxInactiveinterval=0&boardSeq=" + tag + "&pageNum=" + ss + "&national=&searchCodeArrayValue1=TOTAL1&searchCodeArrayValue2=TOTAL2&searchCodeArrayValue3=TOTAL3&searchCodeArrayValue4=TOTAL4&search_keyword=&searchCode1=TOTAL1&searchCode2=TOTAL2&legal_issue=&searchCode3=TOTAL3&searchCode4=TOTAL4&judgment_start_date=&judgment_end_date=&register_start_date=&register_end_date=" )

    b = pd.DataFrame ( s1 )

    # 업로더
     # c

    # 날짜
    d = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(5)' )
    d = pd.DataFrame ( d )
    s1 = []
    time_format = "%Y-%m-%d"
    for tag in d[0] :
        s1.append ( datetime.datetime.strptime ( tag , time_format ) )

    d = pd.DataFrame ( s1 )

    # 조회수
    e = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(4)' )
    e = pd.DataFrame ( e )
    s1 = []
    for tag in e[0] :
        s1.append ( int ( tag ) )

    e = pd.DataFrame ( s1 )
    s1 = []
    for tag in e[0] :
        s1.append ( int ( tag ) )

    e = pd.DataFrame ( s1 )

    # 분류
    f = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(2)' )
    f = pd.DataFrame ( f )

    # 번호
    g = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(1)' )
    g = pd.DataFrame ( g )

    # 헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
    df = []
    df = g
    df = pd.merge ( df , f , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , a , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , b , left_index=True , right_index=True , how='outer' )
    df['업로더'] = '지재권판례'
    df = pd.merge ( df , e , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )
    return df

#1

ss=1
IP_jp = function_IP_judicial_precedent(ss) # 1번

# 2번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[2]').click()
driver.implicitly_wait(3)
ss=2
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 3번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[4]').click()
driver.implicitly_wait(3)
ss=3
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 4번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[5]').click()
driver.implicitly_wait(3)
ss=4
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 5번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[6]').click()
driver.implicitly_wait(3)
ss=5
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 6번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[7]').click()
driver.implicitly_wait(3)
ss=6
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 7번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[8]').click()
driver.implicitly_wait(3)
ss=7
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 8번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[9]').click()
driver.implicitly_wait(3)
ss=8
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 9번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[10]').click()
driver.implicitly_wait(3)
ss=9
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

# 10번
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[11]').click()
driver.implicitly_wait(3)
ss=10
a = function_IP_judicial_precedent(ss)
IP_jp=pd.concat([IP_jp,a])

IP_jp.columns=['번호', '분류', "제목", '링크',"작성자", "조회", "등록일"]
# IP_jp.to_excel('./Text3.xlsx',sheet_name='전체')

# < 지재권 판례 2 - 실무적관점 >

driver.get('https://www.ip-navi.or.kr/board/boardList.navi?boardCode=B00030')
driver.implicitly_wait(3)

# url="https://www.ip-navi.or.kr/board/boardList.navi?boardCode=B00030"
# page = requests.get(url)
# soup = bs(page.text, "html.parser")

page = driver.page_source
soup = bs ( page , "html.parser" )

# 헤드라인
a = []
a=soup.select( '#contents > div.content_area > table > tbody > tr > td.td_tit' )
s1 = []
for tag in a :
    s1.append ( tag.text )

s2 = []
for tag in s1 :
    s2.append (
        tag.replace ( '\n' , '' ).replace ( '\t' , '' ).replace ( '  ' , '' ).replace ( '\xa0' , '' ).replace (
            '\r' , '' ) )

a = pd.DataFrame ( s2 )

# 링크
b = soup.select ( '#contents > div.content_area > table > tbody > tr > td.td_tit > a' )
s1 = []
for tag in b :
    s1.append ( tag["onclick"] )

s1 = pd.DataFrame ( s1 )
v_split = s1[0].str.split ( '(' )
s1 = pd.DataFrame ( v_split.str.get ( 1 ) )
v_split = s1[0].str.split ( ')' )
s1 = pd.DataFrame ( v_split.str.get ( 0 ) )
s2 = []
for tag in s1[0] :
    s2.append ( tag[1 :-1] )

s1 = []
for tag in s2 :
    s1.append (
        "https://www.ip-navi.or.kr/board/viewBoardArticle.navi?boardCode=B00030&getMaxInactiveinterval=0&boardSeq=" + tag + "&pageNum=1&national=&searchCodeArrayValue1=&searchCodeArrayValue2=&searchCodeArrayValue3=&searchCodeArrayValue4=&search_keyword=" )

b = pd.DataFrame ( s1 )

# 업로더
# c

# 날짜
d = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(5)' )
d = pd.DataFrame ( d )
s1 = []
time_format = "%Y-%m-%d"
for tag in d[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

d = pd.DataFrame ( s1 )

# 조회수
e = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(4)' )
e = pd.DataFrame ( e )
s1 = []
for tag in e[0] :
    s1.append ( int ( tag ) )

e = pd.DataFrame ( s1 )
s1 = []
for tag in e[0] :
    s1.append ( int ( tag ) )

e = pd.DataFrame ( s1 )

# 분류
 #f

# 번호
g = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(1)' )
g = pd.DataFrame ( g )

# 헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
df = []
df = g
df['분류'] = '실무적관점'
df = pd.merge ( df , a , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , b , left_index=True , right_index=True , how='outer' )
df['업로더'] = '실무적관점'
df = pd.merge ( df , e , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )

df.columns=['번호', '분류', "제목", '링크',"작성자", "조회", "등록일"]
IP_jp=pd.concat([IP_jp,df])


# < 지재권 판례 3 - 카드뉴스동영상 >

driver.get('https://www.ip-navi.or.kr/board/boardList.navi?boardCode=B00028')
driver.implicitly_wait(3)

# url="https://www.ip-navi.or.kr/board/boardList.navi?boardCode=B00028"
# page = requests.get(url)
# soup = bs(page.text, "html.parser")

page = driver.page_source
soup = bs ( page , "html.parser" )

# 헤드라인
a = []
a=soup.select( '#contents > div.content_area > table > tbody > tr > td.td_tit' )
s1 = []
for tag in a :
    s1.append ( tag.text )

s2 = []
for tag in s1 :
    s2.append (
        tag.replace ( '\n' , '' ).replace ( '\t' , '' ).replace ( '  ' , '' ).replace ( '\xa0' , '' ).replace (
            '\r' , '' ) )

a = pd.DataFrame ( s2 )

# 링크
b = soup.select ( '#contents > div.content_area > table > tbody > tr > td.td_tit > a' )
s1 = []
for tag in b :
    s1.append ( tag["onclick"] )

s1 = pd.DataFrame ( s1 )
v_split = s1[0].str.split ( '(' )
s1 = pd.DataFrame ( v_split.str.get ( 1 ) )
v_split = s1[0].str.split ( ')' )
s1 = pd.DataFrame ( v_split.str.get ( 0 ) )
s2 = []
for tag in s1[0] :
    s2.append ( tag[1 :-1] )

s1 = []
ss=1
ss = str ( ss )
for tag in s2 :
    s1.append (
        "https://www.ip-navi.or.kr/board/viewBoardArticle.navi?boardCode=B00028&getMaxInactiveinterval=0&boardSeq=" + tag + "&pageNum=" +ss+"&national=&searchCodeArrayValue1=&searchCodeArrayValue2=&searchCodeArrayValue3=&searchCodeArrayValue4=&search_keyword=" )

b = pd.DataFrame ( s1 )

# 업로더
# c

# 날짜
d = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(5)' )
d = pd.DataFrame ( d )
s1 = []
time_format = "%Y-%m-%d"
for tag in d[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

d = pd.DataFrame ( s1 )

# 조회수
e = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(4)' )
e = pd.DataFrame ( e )
s1 = []
for tag in e[0] :
    s1.append ( int ( tag ) )

e = pd.DataFrame ( s1 )
s1 = []
for tag in e[0] :
    s1.append ( int ( tag ) )

e = pd.DataFrame ( s1 )

# 분류
 #f

# 번호
g = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(1)' )
g = pd.DataFrame ( g )

# 헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
df = []
df = g
df['분류'] = '카드뉴스동영상'
df = pd.merge ( df , a , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , b , left_index=True , right_index=True , how='outer' )
df['업로더'] = '카드뉴스동영상'
df = pd.merge ( df , e , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )

df.columns=['번호', '분류', "제목", '링크',"작성자", "조회", "등록일"]
IP_jp=pd.concat([IP_jp,df])

## 카드뉴스 다음장
driver.find_element_by_xpath('//*[@id="contents"]/div[3]/a[2]').click()
driver.implicitly_wait(3)

page = driver.page_source
soup = bs ( page , "html.parser" )

# 헤드라인
a = []
a=soup.select( '#contents > div.content_area > table > tbody > tr > td.td_tit' )
s1 = []
for tag in a :
    s1.append ( tag.text )

s2 = []
for tag in s1 :
    s2.append (
        tag.replace ( '\n' , '' ).replace ( '\t' , '' ).replace ( '  ' , '' ).replace ( '\xa0' , '' ).replace (
            '\r' , '' ) )

a = pd.DataFrame ( s2 )

# 링크
b = soup.select ( '#contents > div.content_area > table > tbody > tr > td.td_tit > a' )
s1 = []
for tag in b :
    s1.append ( tag["onclick"] )

s1 = pd.DataFrame ( s1 )
v_split = s1[0].str.split ( '(' )
s1 = pd.DataFrame ( v_split.str.get ( 1 ) )
v_split = s1[0].str.split ( ')' )
s1 = pd.DataFrame ( v_split.str.get ( 0 ) )
s2 = []
for tag in s1[0] :
    s2.append ( tag[1 :-1] )

s1 = []
ss=1
ss = str ( ss )
for tag in s2 :
    s1.append (
        "https://www.ip-navi.or.kr/board/viewBoardArticle.navi?boardCode=B00028&getMaxInactiveinterval=0&boardSeq=" + tag + "&pageNum=" +ss+"&national=&searchCodeArrayValue1=&searchCodeArrayValue2=&searchCodeArrayValue3=&searchCodeArrayValue4=&search_keyword=" )

b = pd.DataFrame ( s1 )

# 업로더
# c

# 날짜
d = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(5)' )
d = pd.DataFrame ( d )
s1 = []
time_format = "%Y-%m-%d"
for tag in d[0] :
    s1.append ( datetime.datetime.strptime ( tag , time_format ) )

d = pd.DataFrame ( s1 )

# 조회수
e = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(4)' )
e = pd.DataFrame ( e )
s1 = []
for tag in e[0] :
    s1.append ( int ( tag ) )

e = pd.DataFrame ( s1 )
s1 = []
for tag in e[0] :
    s1.append ( int ( tag ) )

e = pd.DataFrame ( s1 )

# 분류
 #f

# 번호
g = soup.select ( '#contents > div.content_area > table > tbody > tr > td:nth-child(1)' )
g = pd.DataFrame ( g )

# 헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
df = []
df = g
df['분류'] = '카드뉴스동영상'
df = pd.merge ( df , a , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , b , left_index=True , right_index=True , how='outer' )
df['업로더'] = '카드뉴스동영상'
df = pd.merge ( df , e , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )

df.columns=['번호', '분류', "제목", '링크',"작성자", "조회", "등록일"]
IP_jp=pd.concat([IP_jp,df])

# IP_jp.to_excel('./Text3.xlsx',sheet_name='전체')




# < 저장 ---------------------------------------------------------------------------------  >


# IP_news.to_excel('./Text1.xlsx',sheet_name='전체')
# IP_Desk_news.to_excel('./Text2.xlsx',sheet_name='전체')
# IP_jp.to_excel('./Text2.xlsx',sheet_name='전체')

# End_result.to_excel('./Text3.xlsx',sheet_name='전체')


ct=datetime.datetime.now()
ctstr=ct.strftime("%Y년 %m월 %d일 %H시%M분%S초")

End_result=[]

End_result=pd.concat([IP_news , IP_Desk_news])

End_result=pd.concat([End_result , IP_jp])


End_result=End_result.sort_values('등록일',ascending=False)
IP_news=IP_news.sort_values('등록일',ascending=False)
IP_Desk_news=IP_Desk_news.sort_values('등록일',ascending=False)
IP_jp=IP_jp.sort_values('등록일',ascending=False)


writer=pd.ExcelWriter('C:/작업서류G/작업서류/0.Crawling/IP_NAVI정보('+ctstr+').xlsx', engine='openpyxl')

End_result.to_excel(writer, sheet_name='전체')
IP_news.to_excel(writer, sheet_name='IP_news')
IP_jp.to_excel(writer, sheet_name='IP_jp')
IP_Desk_news.to_excel(writer, sheet_name='IP_Desk_news')

writer.save()
# 엑셀 간격 조정
from openpyxl import load_workbook
wb = load_workbook('C:/작업서류G/작업서류/0.Crawling/IP_NAVI정보('+ctstr+').xlsx')
for tag in wb.sheetnames:
    ws = wb[tag]
    ws.column_dimensions['A'].width=3
    ws.column_dimensions['B'].width=10
    ws.column_dimensions['C'].width=10
    ws.column_dimensions['D'].width=80
    ws.column_dimensions['E'].width=20
    ws.column_dimensions['F'].width=10
    ws.column_dimensions['G'].width=7
    ws.column_dimensions['H'].width=20

wb.save('C:/작업서류G/작업서류/0.Crawling/IP_NAVI정보('+ctstr+').xlsx')

