
import requests
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup as bs
from pandas import DataFrame
from selenium import webdriver
import time
import openpyxl
import datetime
from datetime import timedelta

# < IP실시간뉴스 --------------------------------------------------------------------------------- 시작 1>


options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
driver = webdriver.Chrome(options=options)
driver.implicitly_wait(3)
driver.get('https://www.ip-navi.or.kr/ipnavi/ref/boardList.navi?boardCode=B00034')
time.sleep(2)

def function_IP_news() :

    page = driver.page_source
    soup = bs(page, "html.parser")
    #헤드라인
    a = soup.select('#frm > div > div.board_wrap > table > tbody > tr> td.align_L > a')
    s1=[]
    for tag in a :
        s1.append(tag.text)
    s2=[]
    for tag in s1 :
        s2.append(tag.replace('\n','').replace('\t','').replace('  ','').replace('\xa0','').replace('\r',''))

    a= pd.DataFrame(s2)

    #링크
    b = soup.select('#frm > div > div.board_wrap > table > tbody > tr> td.align_L > a')
    s1=[]
    for tag in b :
        s1.append(tag["href"])

    s1=pd.DataFrame(s1)
    v_split = s1[0].str.split('(')
    s1= pd.DataFrame(v_split.str.get(1))

    v_split = s1[0].str.split(',')
    s2= pd.DataFrame(v_split.str.get(0))
    s3=pd.DataFrame(v_split.str.get(1))
    s4=[]
    for tag in s3[0] :
        s4.append(tag[1:-2])
    s5=[]
    for tag1,tag2 in zip(s2[0],s4):
        s5.append("https://www.ip-navi.or.kr/ipnavi/ref/boardDetail.navi;jsessionid=8360266B59089C27609F57DA66864AD5?boardCode="+tag1[1:-1]+"&boardSeq="+tag2)

    b= pd.DataFrame(s5)

    # Uploader
    s1=[]
    for tag in range(len(a)) :
        s1.append("IP_NAVI_NEWS")

    c= pd.DataFrame(s1)

    # 날짜
    d=soup.select('#frm > div > div.board_wrap > table > tbody > tr > td:nth-child(4)')
    d= pd.DataFrame(d)
    s1=[]
    time_format= "%Y-%m-%d"
    for tag in d[0]:
          s1.append ( datetime.datetime.strptime ( tag , time_format ) )

    d= pd.DataFrame(s1)

    # 조회수
    e=soup.select('#frm > div > div.board_wrap > table > tbody > tr > td:nth-child(3)')
    e= pd.DataFrame(e)

    # 분류
    s1=[]
    for tag in range(len(a)) :
        s1.append("None")

    f= pd.DataFrame(s1)

    #번호
    g=soup.select('#frm > div > div.board_wrap > table > tbody > tr > td:nth-child(1)')
    g= pd.DataFrame(g)

    #헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
    df=[]
    df= pd.merge(g,f, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,a, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,b, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,c, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,e, left_index=True, right_index=True, how='outer')
    df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )
    return df


from selenium.webdriver.common.keys import Keys

# 1번
IP_news=[]
IP_news = function_IP_news() # 1번

# 2번
driver.find_element_by_xpath('//*[@id="divPagination"]/a[1]').send_keys(Keys.ENTER)
time.sleep(0.5)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 3번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[4]').send_keys(Keys.ENTER)
time.sleep(0.5)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 4번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[5]').send_keys(Keys.ENTER)
time.sleep(0.5)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 5번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[6]').send_keys(Keys.ENTER)
time.sleep(0.5)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 6번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[7]').send_keys(Keys.ENTER)
time.sleep(0.5)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 7번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[8]').send_keys(Keys.ENTER)
time.sleep(0.5)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 8번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[9]').send_keys(Keys.ENTER)
time.sleep(0.5)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 9번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[9]').send_keys(Keys.ENTER)
time.sleep(0.5)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 10번

driver.find_element_by_xpath('//*[@id="divPagination"]/a[9]').send_keys(Keys.ENTER)
time.sleep(0.5)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

IP_news.columns=['번호', '분류', "제목", '링크',"작성자", "조회", "등록일"]
# IP_news.to_excel('./Text1.xlsx',sheet_name='전체')




# < 저장 ---------------------------------------------------------------------------------  >


# IP_news.to_excel('./Text1.xlsx',sheet_name='전체')
# IP_Desk_news.to_excel('./Text2.xlsx',sheet_name='전체')
# IP_jp.to_excel('./Text2.xlsx',sheet_name='전체')

# End_result.to_excel('./Text3.xlsx',sheet_name='전체')


ct=datetime.datetime.now()
ctstr=ct.strftime("%Y년 %m월 %d일 %H시%M분%S초")

End_result=IP_news

# End_result=pd.concat([IP_news , IP_Desk_news])
#
# End_result=pd.concat([End_result , IP_jp])


End_result=End_result.sort_values('등록일',ascending=False)
IP_news=IP_news.sort_values('등록일',ascending=False)
# IP_Desk_news=IP_Desk_news.sort_values('등록일',ascending=False)
# IP_jp=IP_jp.sort_values('등록일',ascending=False)


writer=pd.ExcelWriter('C:/작업서류G/작업서류/0.Crawling/IP_NAVI정보('+ctstr+').xlsx', engine='openpyxl')

End_result.to_excel(writer, sheet_name='전체')
IP_news.to_excel(writer, sheet_name='IP_news')
# IP_jp.to_excel(writer, sheet_name='IP_jp')
# IP_Desk_news.to_excel(writer, sheet_name='IP_Desk_news')

writer.save()
# 엑셀 간격 조정
from openpyxl import load_workbook
wb = load_workbook('C:/작업서류G/작업서류/0.Crawling/IP_NAVI정보('+ctstr+').xlsx')
for tag in wb.sheetnames:
    ws = wb[tag]
    ws.column_dimensions['A'].width=3
    ws.column_dimensions['B'].width=10
    ws.column_dimensions['C'].width=10
    ws.column_dimensions['D'].width=80
    ws.column_dimensions['E'].width=20
    ws.column_dimensions['F'].width=10
    ws.column_dimensions['G'].width=7
    ws.column_dimensions['H'].width=20

wb.save('C:/작업서류G/작업서류/0.Crawling/IP_NAVI정보('+ctstr+').xlsx')


# -------------------------------------------------------------------------------------------------------------------------------기업마당

url='C:/작업서류G/작업서류/0.Crawling/'


options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
driver = webdriver.Chrome(options=options)
driver.implicitly_wait(3)
driver.get('https://www.bizinfo.go.kr/web/lay1/bbs/S1T122C128/AS/74/list.do')
time.sleep(1)

def function_기업마당() :
    page = driver.page_source
    soup = bs(page, "html.parser")
    #헤드라인
    a = soup.select('html>body>div:nth-child(1)>section>div:nth-child(3)>form>div:nth-child(3)>div:nth-child(3)>table>tbody>tr>td:nth-child(3)')
    s1=[]
    for tag in a :
        s1.append(tag.text)
    s2=[]
    for tag in s1 :
        s2.append(tag.replace('\n','').replace('\t','').replace('  ','').replace('\xa0','').replace('\r',''))
    a= pd.DataFrame(s2)
    #링크
    b = soup.select('html>body>div:nth-child(1)>section>div:nth-child(3)>form>div:nth-child(3)>div:nth-child(3)>table>tbody>tr>td:nth-child(3)>a')
    s1=[]
    for tag in b :
        s1.append(tag["href"])
    s2=[]
    for tag in s1 :
        s2.append("https://www.bizinfo.go.kr/web/lay1/bbs/S1T122C128/AS/74/"+tag)
    b= pd.DataFrame(s2)
    # Uploader
    c = soup.select('html>body>div:nth-child(1)>section>div:nth-child(3)>form>div:nth-child(3)>div:nth-child(3)>table>tbody>tr>td:nth-child(5)')
    c= pd.DataFrame(c)
    # 날짜
    d = soup.select('#articleSearchForm > div.support_project > div.table_Type_1 >table>tbody>tr>td:nth-child(4)')
    d= pd.DataFrame(d)
    s1=[]
    s2=[]
    s1=pd.DataFrame(d)
    v_split = s1[0].str.split('~')
    s1= list(v_split.str.get(0))
    s2= list(v_split.str.get(1))
    s3=[]
    s4=[]
    for tag in s1 :
        s3.append(str(tag).replace('\n','').replace(' ',''))
    for tag in s2 :
        s4.append(str(tag).replace('\n','').replace(' ',''))
    time_format= "%Y-%m-%d"
    s1=[]
    for tag in s3:
          try:
              s1.append ( datetime.datetime.strptime ( tag , time_format ) )
          except ValueError:
              s1.append(None)

    s2=[]
    for tag in s4:
        try:
            s2.append ( datetime.datetime.strptime ( tag , time_format ) )
        except ValueError :
            s2.append ( None )
    d=pd.DataFrame(s1) # 시작날짜
    e=pd.DataFrame(s2) # 종료날짜
    # 조회수
    f=soup.select('html>body>div:nth-child(1)>section>div:nth-child(3)>form>div:nth-child(3)>div:nth-child(3)>table>tbody>tr>td:nth-child(7)')
    s1 = []
    for tag in f :
        ss=str(tag).replace('<td>','').replace('</td>','')
        s1.append(int(ss))
    f = pd.DataFrame ( s1 )
    # 분류
    g=soup.select('html>body>div:nth-child(1)>section>div:nth-child(3)>form>div:nth-child(3)>div:nth-child(3)>table>tbody>tr>td:nth-child(2)')
    s1 = []
    for tag in g :
        ss=str(tag).replace('<td>','').replace('</td>','').replace(' ','').replace('\n','')
        s1.append(ss)
    g= pd.DataFrame(s1)
    #번호
    h=soup.select('html>body>div:nth-child(1)>section>div:nth-child(3)>form>div:nth-child(3)>div:nth-child(3)>table>tbody>tr>td:nth-child(1)')
    s1=[]
    for tag in h :
        ss=str(tag).replace('<td>','').replace('</td>','')
        s1.append(int(ss))
    h= pd.DataFrame(h)
    s1=[]
    for tag in range(1,len(h)+1):
        s1.append('기업마당')
    i=pd.DataFrame(s1)
    #헤드라인 a, 링크 b, 업로더 c, 시작 날짜 d, 종료 날짜 e,  조회수 f, 지원분야 g, 번호 h , 사이트 i
    df=[]
    df= pd.merge(i,h, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,g, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,a, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,b, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,c, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,d, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,e, left_index=True, right_index=True, how='outer')
    df = pd.merge ( df , f , left_index=True , right_index=True , how='outer' )
    df.columns=["사이트", "번호","지원분야" ,"지원사업명", "링크","소관부처", "시작", "종료", "조회수"]
    return df

# 1번 ~ 2번

공고_기업마당 = function_기업마당() # 1번
driver.find_element_by_xpath('//*[@id="container"]/div[3]/div[2]/a[3]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_기업마당()
공고_기업마당=pd.concat([공고_기업마당,a]) #2번 합치기

# 3번
driver.find_element_by_xpath('//*[@id="container"]/div[3]/div[2]/a[5]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_기업마당()
공고_기업마당=pd.concat([공고_기업마당,a])
# 4번
driver.find_element_by_xpath('//*[@id="container"]/div[3]/div[2]/a[6]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_기업마당()
공고_기업마당=pd.concat([공고_기업마당,a])
# 5번
driver.find_element_by_xpath('//*[@id="container"]/div[3]/div[2]/a[7]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_기업마당()
공고_기업마당=pd.concat([공고_기업마당,a])
# 6번
driver.find_element_by_xpath('//*[@id="container"]/div[3]/div[2]/a[8]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_기업마당()
공고_기업마당=pd.concat([공고_기업마당,a])
# 7번
driver.find_element_by_xpath('//*[@id="container"]/div[3]/div[2]/a[9]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_기업마당()
공고_기업마당=pd.concat([공고_기업마당,a])
# 8번
driver.find_element_by_xpath('//*[@id="container"]/div[3]/div[2]/a[10]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_기업마당()
공고_기업마당=pd.concat([공고_기업마당,a])
# 9번
driver.find_element_by_xpath('//*[@id="container"]/div[3]/div[2]/a[11]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_기업마당()
공고_기업마당=pd.concat([공고_기업마당,a])
# 10번
driver.find_element_by_xpath('//*[@id="container"]/div[3]/div[2]/a[12]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_기업마당()
공고_기업마당=pd.concat([공고_기업마당,a])

공고_기업마당.columns=["사이트", "번호","지원분야" ,"지원사업명", "링크","소관부처", "시작", "종료", "조회수"]

# 공고_기업마당.to_excel(url+'기업마당'+tm+').xlsx', sheet_name='Sheet1')
# --------------------------------------------------------------------------------------------------------------------기업마당
# --------------------------------------------------------------------------------------------------------------------K-스타트업
page=[]
soup=[]
a=[]

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
driver = webdriver.Chrome(options=options)
driver.implicitly_wait(3)
driver.get('https://www.k-startup.go.kr/web/contents/bizpbanc-ongoing.do')

def function_kstar() :
    page = driver.page_source
    soup = bs(page, "html.parser")
    #헤드라인
    a = soup.select('#bizPbancList > ul > li > div > div.right > div.middle > a > div > p')
    s1=[]
    for tag in a :
        s1.append(tag.text)
    a= pd.DataFrame(s1)
    #링크
    b = soup.select('#bizPbancList > ul > li > div > div.right > div.middle > a')
    s1=[]
    for tag in b :
        s1.append(tag["href"])
    s2=[]
    for tag in s1 :
        s2.append(tag.replace('javascript:go_view(','').replace(');',''))
    s3=[]
    for tag in s2 :
        s3.append("https://www.k-startup.go.kr/web/contents/bizpbanc-ongoing.do?schM=view&pbancSn="+tag+'&page=1&schStr=regist&pbancEndYn=N')
    b= pd.DataFrame(s3)
    # Uploader
    c = soup.select('#bizPbancList > ul > li > div > div.right > div.bottom > span:nth-child(2)')
    s1=[]
    for tag in c :
        s1.append(tag.text.replace(' ',''))
    c= pd.DataFrame(s1)
    # 시작 날짜
    d = soup.select('#bizPbancList > ul > li > div > div.right > div.bottom > span:nth-child(3)')
    s1=[]
    for tag in d :
        s1.append(tag.text.replace(' ','').replace('등록일자',''))
    time_format= "%Y-%m-%d"
    s2=[]
    for tag in s1:
          s2.append ( datetime.datetime.strptime ( tag , time_format ) )
    d=pd.DataFrame(s2) # 시작날짜
    # 마감 일자
    e = soup.select('#bizPbancList > ul > li > div > div.right > div.bottom > span:nth-child(4)')
    s1=[]
    for tag in e :
        s1.append(tag.text.replace(' ','').replace('마감일자',''))
    time_format= "%Y-%m-%d"
    s2=[]
    for tag in s1:
          s2.append ( datetime.datetime.strptime ( tag , time_format ) )
    e=pd.DataFrame(s2) # 마감일자
    # 조회수
    f=soup.select('#bizPbancList > ul > li > div > div.right > div.bottom > span:nth-child(5)')
    s1=[]
    for tag in f :
        ss=tag.text.replace(' ','').replace('조회','').replace(',','')
        s1.append(int(ss))
    f = pd.DataFrame ( s1 )
    # 분류
    g=soup.select('#bizPbancList > ul > li > div > div.right > div.top > span:nth-child(1)')
    s1 = []
    for tag in g :
        ss=tag.text.replace('<td>','').replace('</td>','').replace(' ','').replace('\n','').replace('\t','')
        s1.append(ss)
    g= pd.DataFrame(s1)
    #번호
    h=soup.select('#bizPbancList > ul > li > div > div.right > div.top > span:nth-child(2)')
    s1 = []
    for tag in h :
        ss=tag.text.replace('<td>','').replace('</td>','').replace(' ','').replace('\n','').replace('\t','')
        s1.append(ss)
    h= pd.DataFrame(h)
    s1=[]
    for tag in range(1,len(h)+1):
        s1.append('K-스타트업')
    i=pd.DataFrame(s1)
    #헤드라인 a, 링크 b, 업로더 c, 시작 날짜 d, 종료 날짜 e,  조회수 f, 지원분야 g, 번호 h , 사이트 i
    df=[]
    df= pd.merge(i,h, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,g, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,a, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,b, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,c, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,d, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,e, left_index=True, right_index=True, how='outer')
    df = pd.merge ( df , f , left_index=True , right_index=True , how='outer' )
    df.columns=["사이트", "번호","지원분야" ,"지원사업명", "링크","소관부처", "시작", "종료", "조회수"]
    return df

# 1번 ~ 2번

공고_kstar = function_kstar() # 1번
driver.find_element_by_xpath('//*[@id="bizPbancList"]/div/a[2]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_kstar()
공고_kstar=pd.concat([공고_kstar,a]) #2번 합치기

# 3번
driver.find_element_by_xpath('//*[@id="bizPbancList"]/div/a[3]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_kstar()
공고_kstar=pd.concat([공고_kstar,a])
# 4번
driver.find_element_by_xpath('//*[@id="bizPbancList"]/div/a[4]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_kstar()
공고_kstar=pd.concat([공고_kstar,a])
# 5번
driver.find_element_by_xpath('//*[@id="bizPbancList"]/div/a[5]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_kstar()
공고_kstar=pd.concat([공고_kstar,a])
# 6번
driver.find_element_by_xpath('//*[@id="bizPbancList"]/div/a[6]').click()
driver.implicitly_wait(3)
time.sleep(1)
a = function_kstar()
공고_kstar=pd.concat([공고_kstar,a])

공고_kstar.columns=["사이트", "번호","지원분야" ,"지원사업명", "링크","소관부처", "시작", "종료", "조회수"]

# 공고_kstar.to_excel(url+'K-star공고('+tm+').xlsx', sheet_name='Sheet1')

# --------------------------------------------------------------------------------------------------------------------K-스타트업
# --------------------------------------------------------------------------------------------------------------------경기 TP

url1="https://pms.gtp.or.kr/web/business/webBusinessList.do?"
page = requests.get(url1)
soup = bs(page.text, "html.parser")

# 헤드라인
a = soup.select ( '#container > div.content > div.inner > table > tbody > tr > td.subject > a' )
s1 = []
for tag in a :
    s1.append ( tag['title'] )

a = pd.DataFrame ( s1 )
# 링크
b = soup.select ( '#container > div.content > div.inner > table > tbody > tr > td.subject > a' )
s1 = []
for tag in b :
    s1.append ( tag["onclick"] )

s2 = []
for tag in s1 :
    s2.append (tag.replace('fn_goView(','').replace('); return','').replace('false','').replace(';','').replace(' ','').replace('\'',''))

s3 = []
for tag in s2 :
    s3.append (
        "https://pms.gtp.or.kr/web/business/webBusinessView.do?b_idx=" + tag )

b = pd.DataFrame ( s3 )
# Uploader
c = soup.select ( '#container > div.content > div.inner > table > tbody > tr > td:nth-child(5)' )
s1 = []
for tag in c :
    s1.append ( tag.text.replace ( ' ' , '' ) )

c = pd.DataFrame ( s1 )
# 시작 날짜
d = soup.select ( '#container > div.content > div.inner > table > tbody > tr > td.last' )
s1 = []
for tag in d :
    s1.append ( tag.text.replace ( ' ' , '' ).replace ( '등록일자' , '' ).replace('\t','').replace('\n','').replace('\r','') )

s2=[]
s3=[]
for tag in s1 :
    if tag=='마감' :
        s3.append(str(tag))
    if tag!='마감' :
        s2.append(str(tag))

ss=[]
s4=[]
s5=[]
for tag in s2 :
    ss=tag.split('~')
    s4.append(ss[0])
    s5.append(ss[1])

time_format = "%Y-%m-%d"
s6 = []
s7 = []
for tag in s4 :
    s6.append ( datetime.datetime.strptime ( tag , time_format ) )

for tag in s5 :
    s7.append ( datetime.datetime.strptime ( tag , time_format ) )

ct=datetime.datetime.now()
ct1= ct  - timedelta(weeks=1)
s8=[]
for tag in range(0,len(s3)):
    s8.append(ct1)


d = pd.DataFrame ( s6+s8 )  # 시작날짜
e=  pd.DataFrame ( s7+s3 )  # 마감날짜

# 조회수
s1 = []
for tag in range(0,len(e)) :
    s1.append ( int('0') )

f = pd.DataFrame ( s1 )

# 분류
g = soup.select ( '#container > div.content > div.inner > table > tbody > tr > td:nth-child(3)' )
s1 = []
for tag in g :
    ss = tag.text.replace ( '<td>' , '' ).replace ( '</td>' , '' ).replace ( ' ' , '' ).replace ( '\n' , '' ).replace (
        '\t' , '' )
    s1.append ( ss )

g = pd.DataFrame ( s1 )

# 번호
h = soup.select ( '#container > div.content > div.inner > table > tbody > tr > td:nth-child(1)' )
s1 = []
for tag in h :
    ss = tag.text.replace ( '<td>' , '' ).replace ( '</td>' , '' ).replace ( ' ' , '' ).replace ( '\n' , '' ).replace (
        '\t' , '' )
    s1.append ( ss )

h = pd.DataFrame ( h )
s1 = []
for tag in range ( 1 , len ( h ) + 1 ) :
    s1.append ( '경기TP' )

i = pd.DataFrame ( s1 )

# 헤드라인 a, 링크 b, 업로더 c, 시작 날짜 d, 종료 날짜 e,  조회수 f, 지원분야 g, 번호 h , 사이트 i
df = []
df = pd.merge ( i , h , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , g , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , a , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , b , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , c , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , e , left_index=True , right_index=True , how='outer' )
df = pd.merge ( df , f , left_index=True , right_index=True , how='outer' )
df.columns = ["사이트" , "번호" , "지원분야" , "지원사업명" , "링크" , "소관부처" , "시작" , "종료" , "조회수"]

공고_경기tp= df.copy()

# < 저장 ---------------------------------------------------------------------------------  >

ct=datetime.datetime.now()
ctstr=ct.strftime("%Y년 %m월 %d일 %H시%M분%S초")

End_result=[]
End_result=pd.concat([공고_기업마당,공고_kstar,공고_경기tp])

공고_기업마당=공고_기업마당.sort_values('시작',ascending=False)
공고_kstar=공고_kstar.sort_values('시작',ascending=False)
공고_경기tp=공고_경기tp.sort_values('번호',ascending=False)
End_result=End_result.sort_values('시작',ascending=False)

writer=pd.ExcelWriter(url+'기업마당+K스타트업+경기TP('+ctstr+').xlsx', engine='openpyxl')

End_result.to_excel(writer, sheet_name='전체')
공고_기업마당.to_excel(writer, sheet_name='공고_기업마당')
공고_kstar.to_excel(writer, sheet_name='공고_kstar')
공고_경기tp.to_excel(writer, sheet_name='공고_경기tp')

writer.save()
# 엑셀 간격 조정
from openpyxl import load_workbook
wb = load_workbook(url+'기업마당+K스타트업+경기TP('+ctstr+').xlsx')
for tag in wb.sheetnames:
    ws = wb[tag]
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 80
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 7

wb.save(url+'기업마당+K스타트업+경기TP('+ctstr+').xlsx')


