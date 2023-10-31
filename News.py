
import requests
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup as bs
from pandas import DataFrame
from selenium import webdriver
import time
import openpyxl
import datetime
from datetime import timedelta

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

# < IP실시간뉴스 --------------------------------------------------------------------------------- 시작 1>



chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=chrome_options)
driver.implicitly_wait(3)
driver.get('https://www.ip-navi.or.kr/ipnavi/ref/boardList.navi?boardCode=B00034')
time.sleep(1)

def function_IP_news() :
    page = driver.page_source
    soup = bs(page, "html.parser")
    #헤드라인
    a = soup.select('#frm > div > div.board_wrap > table > tbody > tr > td.align_L > a')
    s1=[]
    for tag in a :
        s1.append(tag.text)
    s2=[]
    for tag in s1 :
        s2.append(tag.replace('\n','').replace('\t','').replace('  ','').replace('\xa0','').replace('\r',''))
    a= pd.DataFrame(s2)
    a.rename(columns={0: 'a'}, inplace=True)
    #링크
    b = soup.select('#frm > div > div.board_wrap > table > tbody > tr > td.align_L > a')
    s1=[]
    for tag in b :
        s1.append(tag["href"])

    s1=pd.DataFrame(s1)
    v_split = s1[0].str.split('(')
    s1= pd.DataFrame(v_split.str.get(1))

    v_split = s1[0].str.split(',')
    s1= pd.DataFrame(v_split.str.get(1))
    s2=[]
    for tag in s1[0] :
        s2.append("https://www.ip-navi.or.kr/ipnavi/ref/boardDetail.navi?boardCode=B00034&boardSeq="+tag[1:-2])

    b= pd.DataFrame(s2)
    b.rename(columns={0: 'b'}, inplace=True)
    # Uploader
    s1=[]
    for tag in range(len(a)) :
        s1.append("IP_NAVI_NEWS")

    c= pd.DataFrame(s1)
    c.rename(columns={0: 'c'}, inplace=True)
    # 날짜
    d=soup.select('#frm > div > div.board_wrap > table > tbody > tr > td:nth-child(4)')
    d= pd.DataFrame(d)
    s1=[]
    time_format= "%Y-%m-%d"
    for tag in d[0]:
          s1.append ( datetime.datetime.strptime ( tag , time_format ) )

    d= pd.DataFrame(s1)
    d.rename(columns={0: 'd'}, inplace=True)
    # 조회수
    e=soup.select('#frm > div > div.board_wrap > table > tbody > tr > td:nth-child(3)')
    e= pd.DataFrame(e)
    s1 = []
    for tag in e[0]:
        s1.append(int(tag))
    e = pd.DataFrame(s1)
    e.rename(columns={0: 'e'}, inplace=True)

    # 분류
    s1=[]
    for tag in range(len(a)) :
        s1.append("None")

    f= pd.DataFrame(s1)
    f.rename(columns={0: 'f'}, inplace=True)
    #번호
    g=soup.select('#frm > div > div.board_wrap > table > tbody > tr > td:nth-child(1)')
    g= pd.DataFrame(g)
    g.rename(columns={0: 'g'}, inplace=True)
    #헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
    df=[]
    df= pd.merge(g,f, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,a, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,b, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,c, left_index=True, right_index=True, how='outer')
    df= pd.merge(df,e, left_index=True, right_index=True, how='outer')
    df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )
    return df


# 1번
IP_news=[]
IP_news = function_IP_news() # 1번

# 2번
element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[1]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(10)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 3번
element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[4]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(10)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 4번

element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[5]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(10)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 5번

element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[6]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(10)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 6번

element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[7]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(10)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 7번

element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[8]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(10)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 8번

element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[9]')
driver.execute_script("arguments[0].click();", element)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 9번

element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[9]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(10)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

# 10번

element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[9]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(10)
a = function_IP_news()
IP_news=pd.concat([IP_news,a])

IP_news.columns=['번호', '분류', "제목", '링크',"작성자", "조회", "등록일"]
# IP_news.to_excel('./Text1.xlsx',sheet_name='전체')

# < IP_Desk_news --------------------------------------------------------------------------------- 종료 2>


chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=chrome_options)
driver.implicitly_wait(10)
driver.get('http://www.ip-navi.or.kr/ipnavi/ref/boardList.navi?boardCode=B00018')
time.sleep(1)


def function_IP_Desk_news() :
    page = driver.page_source
    soup = bs ( page , "html.parser" )

    # 헤드라인
    a = []
    a = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr> td.align_L' )
    s1 = []
    for tag in a :
        s1.append ( tag.text )

    s2 = []
    for tag in s1 :
        s2.append (
            tag.replace ( '\n' , '' ).replace ( '\t' , '' ).replace ( '  ' , '' ).replace ( '\xa0' , '' ).replace (
                '\r' , '' ) )

    a = pd.DataFrame ( s2 )
    a.rename(columns={0: 'a'}, inplace=True)
    # 링크
    s1 = [element['onclick'] for element in soup.find_all('tr', attrs={'onclick': True})]
    s1 = pd.DataFrame ( s1 )
    v_split = s1[0].str.split ( ',' )
    s1 = pd.DataFrame ( v_split.str.get ( 1 ) )
    v_split = s1[0].str.split ( ')' )
    s1 = pd.DataFrame ( v_split.str.get ( 0 ) )
    s2 = []
    for tag in s1[0] :
        s2.append ( tag[2 :-1] )
    s1 = []
    for tag in s2 :
        s1.append (
            "https://www.ip-navi.or.kr/ipnavi/ref/boardDetail.navi?boardCode=B00018&boardSeq=" + tag +"&currentPageNo=1" )

    b = pd.DataFrame ( s1 )
    b.rename(columns={0: 'b'}, inplace=True)

    # Uploader
    s1 = []
    for tag in range ( len ( a ) ) :
        s1.append ( "IP_DESK_NEWS" )
    c = pd.DataFrame ( s1 )
    c.rename(columns={0: 'c'}, inplace=True)
    # 날짜
    d = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr > td:nth-child(4)' )
    d = pd.DataFrame ( d )
    s1 = []
    time_format = "%Y-%m-%d"
    for tag in d[0] :
        s1.append ( datetime.datetime.strptime ( tag , time_format ) )

    d = pd.DataFrame ( s1 )
    d.rename(columns={0: 'd'}, inplace=True)
    # 조회수
    e = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr > td:nth-child(3)' )
    e = pd.DataFrame ( e )
    s1 = []
    for tag in e[0] :
        s1.append ( int ( tag ) )

    e = pd.DataFrame ( s1 )
    e.rename(columns={0: 'e'}, inplace=True)
    # 분류
    s1 = []
    for tag in range ( len ( a ) ) :
        s1.append ( "NONE" )
    f = pd.DataFrame ( s1 )
    f.rename(columns={0: 'f'}, inplace=True)
    # 번호
    g = soup.select ( '#frm > div > div.board_wrap > table > tbody > tr > td:nth-child(1)' )
    g = pd.DataFrame ( g )
    g.rename(columns={0: 'g'}, inplace=True)
    # 헤드라인 a, 링크 b, 업로더 c, 날짜 d, 조회수 e, 분류 f, 번호 g
    df = []
    df = g
    df['분류'] = 'IP_DESK_지재권뉴스'
    df = pd.merge ( df , a , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , b , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , c , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , e , left_index=True , right_index=True , how='outer' )
    df = pd.merge ( df , d , left_index=True , right_index=True , how='outer' )
    return df


#1
IP_Desk_news=[]
IP_Desk_news = function_IP_Desk_news() # 1번

# 2번
element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[1]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(10)
a = function_IP_Desk_news()
IP_Desk_news=pd.concat([IP_Desk_news,a])

# 3번
element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[4]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(10)
a = function_IP_Desk_news()
IP_Desk_news=pd.concat([IP_Desk_news,a])

# 4번
element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[5]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(10)
a = function_IP_Desk_news()
IP_Desk_news=pd.concat([IP_Desk_news,a])

# 5번
element = driver.find_element(By.XPATH,'//*[@id="divPagination"]/a[6]')
driver.execute_script("arguments[0].click();", element)
driver.implicitly_wait(10)
a = function_IP_Desk_news()
IP_Desk_news=pd.concat([IP_Desk_news,a])


IP_Desk_news.columns=['번호', '분류', "제목", '링크',"작성자", "조회", "등록일"]

# IP_Desk_news.to_excel('./Text2.xlsx',sheet_name='전체')

# < 저장 ------------------------------------------------------------------------------------------------------ 저장 >

ct=datetime.datetime.now()
ctstr=ct.strftime("%Y년 %m월 %d일 %H시%M분%S초")

writer=pd.ExcelWriter('C:/작업서류SG/작업서류/0.Crawling/IP_NEWS('+ctstr+').xlsx', engine='openpyxl')
IP_news.to_excel(writer, sheet_name='IP_news')
IP_Desk_news.to_excel(writer, sheet_name='IP_Desk_news')
writer.close()

# 엑셀 간격 조정
from openpyxl import load_workbook
wb = load_workbook('C:/작업서류SG/작업서류/0.Crawling/IP_NEWS('+ctstr+').xlsx')
for tag in wb.sheetnames:
    ws = wb[tag]
    ws.column_dimensions['A'].width=3
    ws.column_dimensions['B'].width=10
    ws.column_dimensions['C'].width=10
    ws.column_dimensions['D'].width=80
    ws.column_dimensions['E'].width=20
    ws.column_dimensions['F'].width=20
    ws.column_dimensions['G'].width=5
    ws.column_dimensions['H'].width=25

wb.save('C:/작업서류SG/작업서류/0.Crawling/IP_NEWS('+ctstr+').xlsx')
